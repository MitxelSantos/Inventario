#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
INVENTORY MANAGER - Sistema de Inventario Tecnológico
=========================================================
Hospital Regional Alfonso Jaramillo Salazar

VERSIÓN: 1.0
FECHA: Enero 2026
"""

import tkinter as tk
from tkinter import messagebox, filedialog
from tkcalendar import DateEntry

import customtkinter as ctk
import platform
import socket
import subprocess
import os
import re
import threading

from datetime import datetime
from pathlib import Path

# Configurar tema CustomTkinter
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("green")

# Importar configuración
try:
    from config_listas import *
except ImportError:
    messagebox.showerror("Error", "No se encontró config_listas.py\nAsegúrate de tener ambos archivos en la misma carpeta")
    exit(1)

# Librerías opcionales
try:
    import openpyxl
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    messagebox.showwarning("Advertencia", "openpyxl no instalado. Ejecuta:\npip install openpyxl")

try:
    import psutil
    HAS_PSUTIL = True
except ImportError:
    HAS_PSUTIL = False

try:
    import wmi
    HAS_WMI = True
except ImportError:
    HAS_WMI = False
    print("WMI no disponible - Detección de hardware limitada")

try:
    import winreg
    HAS_WINREG = True
except ImportError:
    HAS_WINREG = False

# PIL para cargar imágenes (logo)
try:
    from PIL import Image, ImageDraw
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("PIL/Pillow no disponible - Logo no se mostrará")


# ============================================================================
# COLORES INSTITUCIONALES
# ============================================================================

COLOR_VERDE_HOSPITAL = "#3C8B0E"
COLOR_AZUL_HOSPITAL = "#3C8B0E"
COLOR_NARANJA = "#F4B183"
COLOR_FONDO = "#F5F5F5"
COLOR_ERROR = "#DC3545"

# ============================================================================
# 1. CLASE TOOLTIP
# ============================================================================

class ToolTip:
    """
    Clase para mostrar tooltips al pasar el mouse sobre un widget.
    """
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)
    
    def show_tooltip(self, event=None):
        if self.tooltip:
            return
        
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")
        
        # Frame con borde
        frame = tk.Frame(self.tooltip, bg="#FFFFCC", relief=tk.SOLID, borderwidth=1)
        frame.pack()
        
        # Label con texto (máximo 600px de ancho)
        label = tk.Label(
            frame, 
            text=self.text, 
            bg="#FFFFCC",
            fg="#000000",
            font=("Segoe UI", 10),
            wraplength=600,
            justify=tk.LEFT,
            padx=10,
            pady=8
        )
        label.pack()
    
    def hide_tooltip(self, event=None):
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None


# ============================================================================
# FUNCIONES DE DETECCIÓN
# ============================================================================

def detect_hardware_wmi():
    """
    Detectar hardware usando WMI.
    """
    info = {
        'marca': 'No detectado',
        'modelo': 'No detectado',
        'serial': 'No detectado',
        # DISCO 1 (Primario)
        'disco1_capacidad': 'No detectado',  
        'disco1_tipo': 'No detectado',      
        'disco1_serial': 'No detectado',
        'disco1_marca': 'No detectado',
        'disco1_modelo': 'No detectado',
        # DISCO 2 (Secundario)
        'disco2_capacidad': 'No tiene',
        'disco2_tipo': 'No tiene',
        'disco2_serial': 'No tiene',
        'disco2_marca': 'No tiene',
        'disco2_modelo': 'No tiene'
    }
    
    if not HAS_WMI:
        return info
    
    try:
        # Inicializar COM
        try:
            import pythoncom
            pythoncom.CoInitialize()
        except:
            pass
        
        c = wmi.WMI()
        
        # ===== INFORMACIÓN DEL SISTEMA =====
        for system in c.Win32_ComputerSystem():
            info['marca'] = system.Manufacturer or 'No detectado'
            info['modelo'] = system.Model or 'No detectado'
        
        # ===== SERIAL DEL EQUIPO =====
        serial_found = False
        serials_invalidos = ['default string', 'to be filled by o.e.m.', 
                            'system serial number', 'base board serial number', 
                            'chassis serial number', '']
        
        # Intentar BIOS primero
        for bios in c.Win32_BIOS():
            serial = (bios.SerialNumber or '').strip()
            if serial and serial.lower() not in serials_invalidos:
                info['serial'] = serial
                serial_found = True
                break
        
        # Intentar BaseBoard
        if not serial_found:
            for board in c.Win32_BaseBoard():
                serial = (board.SerialNumber or '').strip()
                if serial and serial.lower() not in serials_invalidos:
                    info['serial'] = f"MB-{serial}"
                    serial_found = True
                    break
        
        # Intentar ComputerSystemProduct
        if not serial_found:
            for product in c.Win32_ComputerSystemProduct():
                serial = (product.IdentifyingNumber or '').strip()
                if serial and serial.lower() not in serials_invalidos:
                    info['serial'] = serial
                    serial_found = True
                    break
        
        if not serial_found:
            info['serial'] = "No detectado (PC genérico/armado)"
        
        # ===== DISCOS FÍSICOS =====
        disks = list(c.Win32_DiskDrive())
        
        # DISCO 1 (PRIMARIO)
        if len(disks) > 0:
            disk1 = disks[0]
            
            # Capacidad en GB
            try:
                size_bytes = int(disk1.Size) if disk1.Size else 0
                size_gb = round(size_bytes / (1024**3))
                info['disco1_capacidad'] = str(size_gb) 
            except:
                info['disco1_capacidad'] = 'No detectado'
            
            # Tipo (SSD o HDD)
            media_type = disk1.MediaType or ''
            if 'SSD' in media_type.upper() or 'Solid State' in media_type:
                info['disco1_tipo'] = 'SSD'
            else:
                info['disco1_tipo'] = 'HDD'  
            
            # Serial
            serial_disk = (disk1.SerialNumber or '').strip()
            info['disco1_serial'] = serial_disk if serial_disk else 'No detectado'
            
            # Marca
            marca_disk = (disk1.Manufacturer or '').strip()
            if marca_disk and marca_disk.lower() not in ['(standard disk drives)', '']:
                info['disco1_marca'] = marca_disk
            else:
                info['disco1_marca'] = 'No detectado'
            
            # Modelo
            modelo_disk = (disk1.Model or '').strip()
            info['disco1_modelo'] = modelo_disk if modelo_disk else 'No detectado'
        
        # DISCO 2 (SECUNDARIO)
        if len(disks) > 1:
            disk2 = disks[1]
            
            try:
                size_bytes = int(disk2.Size) if disk2.Size else 0
                size_gb = round(size_bytes / (1024**3))
                info['disco2_capacidad'] = str(size_gb)
            except:
                info['disco2_capacidad'] = 'Detectado'
            
            media_type = disk2.MediaType or ''
            if 'SSD' in media_type.upper() or 'Solid State' in media_type:
                info['disco2_tipo'] = 'SSD'
            else:
                info['disco2_tipo'] = 'HDD'
            
            serial_disk2 = (disk2.SerialNumber or '').strip()
            info['disco2_serial'] = serial_disk2 if serial_disk2 else 'No detectado'
            
            marca_disk2 = (disk2.Manufacturer or '').strip()
            if marca_disk2 and marca_disk2.lower() not in ['(standard disk drives)', '']:
                info['disco2_marca'] = marca_disk2
            else:
                info['disco2_marca'] = 'No detectado'
            
            modelo_disk2 = (disk2.Model or '').strip()
            info['disco2_modelo'] = modelo_disk2 if modelo_disk2 else 'No detectado'
    
    except Exception as e:
        print(f"Error WMI: {e}")
    
    return info


def detect_office_version():
    """Detectar versión de Office: Busca ejecutables incluso sin licencia."""
    if not HAS_WINREG:
        return "No detectado", "No detectado"
    
    try:
        # ESTRATEGIA 1: Buscar en InstallRoot (instalación completa licenciada)
        key_paths = [
            r"SOFTWARE\Microsoft\Office\16.0\Common\InstallRoot",  # Office 2016/2019/365
            r"SOFTWARE\Microsoft\Office\15.0\Common\InstallRoot",  # Office 2013
            r"SOFTWARE\Microsoft\Office\14.0\Common\InstallRoot",  # Office 2010
        ]
        
        for key_path in key_paths:
            try:
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path)
                path = winreg.QueryValueEx(key, "Path")[0]
                winreg.CloseKey(key)
                
                if "16.0" in key_path:
                    version = "Office 2016/2019/365"
                elif "15.0" in key_path:
                    version = "Office 2013"
                elif "14.0" in key_path:
                    version = "Office 2010"
                else:
                    version = "Detectado"
                
                licencia = "Retail/Volume"
                return version, licencia
            except:
                continue
        
        # ESTRATEGIA 2: Buscar ejecutables de Office (incluso sin licencia completa)
        office_paths = [
            (r"C:\Program Files\Microsoft Office\root\Office16\WINWORD.EXE", "Office 2016/2019/365"),
            (r"C:\Program Files (x86)\Microsoft Office\root\Office16\WINWORD.EXE", "Office 2016/2019/365"),
            (r"C:\Program Files\Microsoft Office\Office16\WINWORD.EXE", "Office 2016/2019/365"),
            (r"C:\Program Files (x86)\Microsoft Office\Office16\WINWORD.EXE", "Office 2016/2019/365"),
            (r"C:\Program Files\Microsoft Office\Office15\WINWORD.EXE", "Office 2013"),
            (r"C:\Program Files (x86)\Microsoft Office\Office15\WINWORD.EXE", "Office 2013"),
            (r"C:\Program Files\Microsoft Office\Office14\WINWORD.EXE", "Office 2010"),
            (r"C:\Program Files (x86)\Microsoft Office\Office14\WINWORD.EXE", "Office 2010"),
        ]
        
        for path, version in office_paths:
            if os.path.exists(path):
                return version, "Instalado (verificar licencia)"
        
        # ESTRATEGIA 3: Buscar en registro de desinstalación
        try:
            for hive in [winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER]:
                try:
                    key = winreg.OpenKey(hive, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall")
                    for i in range(winreg.QueryInfoKey(key)[0]):
                        try:
                            subkey_name = winreg.EnumKey(key, i)
                            if 'Office' in subkey_name or 'Microsoft 365' in subkey_name:
                                subkey = winreg.OpenKey(key, subkey_name)
                                try:
                                    display_name = winreg.QueryValueEx(subkey, "DisplayName")[0]
                                    if 'Office' in display_name or 'Microsoft 365' in display_name:
                                        winreg.CloseKey(subkey)
                                        winreg.CloseKey(key)
                                        return display_name, "Instalado (verificar licencia)"
                                except:
                                    pass
                                winreg.CloseKey(subkey)
                        except:
                            continue
                    winreg.CloseKey(key)
                except:
                    continue
        except:
            pass
        
        return "No instalado", "N/A"
    
    except Exception as e:
        return "No detectado", "No detectado"


def detect_office_apps():
    """Detectar si Teams y Outlook están instalados."""
    teams = "No"
    outlook = "No"
    
    # Rutas comunes de Teams
    teams_paths = [
        r"C:\Users\{}\AppData\Local\Microsoft\Teams\current\Teams.exe",
        r"C:\Program Files\Microsoft\Teams\current\Teams.exe",
        r"C:\Program Files (x86)\Microsoft\Teams\current\Teams.exe"
    ]
    
    username = os.environ.get('USERNAME', '')
    for path in teams_paths:
        full_path = path.format(username)
        if os.path.exists(full_path):
            teams = "Sí"
            break
    
    # Rutas comunes de Outlook
    outlook_paths = [
        r"C:\Program Files\Microsoft Office\root\Office16\OUTLOOK.EXE",
        r"C:\Program Files (x86)\Microsoft Office\root\Office16\OUTLOOK.EXE",
        r"C:\Program Files\Microsoft Office\Office16\OUTLOOK.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office16\OUTLOOK.EXE",
    ]
    
    for path in outlook_paths:
        if os.path.exists(path):
            outlook = "Sí"
            break
    
    return teams, outlook


def detect_windows_license():
    """Detectar información de licencia de Windows."""
    licencia_info = {
        'tipo': 'No detectado',
        'key': 'No detectado',
        'estado': 'No detectado'
    }
    
    try:
        # Ejecutar slmgr para obtener info de licencia
        result = subprocess.run(
            ['cscript', '//nologo', r'C:\Windows\System32\slmgr.vbs', '/dli'],
            capture_output=True,
            text=True,
            timeout=10
        )
        
        output = result.stdout
        
        # Parsear tipo de licencia
        if 'OEM' in output:
            licencia_info['tipo'] = 'OEM'
        elif 'Retail' in output:
            licencia_info['tipo'] = 'Retail'
        elif 'Volume' in output:
            licencia_info['tipo'] = 'Volume'
        else:
            licencia_info['tipo'] = 'Detectado'
        
        # Estado
        if 'Licensed' in output or 'Licenciado' in output:
            licencia_info['estado'] = 'Activado'
        else:
            licencia_info['estado'] = 'No activado'
        
        # Obtener últimos 5 dígitos de la key
        key_result = subprocess.run(
            ['cscript', '//nologo', r'C:\Windows\System32\slmgr.vbs', '/dli'],
            capture_output=True,
            text=True,
            timeout=10
        )
        
        key_output = key_result.stdout
        # Buscar patrón de product key (últimos 5)
        key_match = re.search(r'([A-Z0-9]{5})$', key_output, re.MULTILINE)
        if key_match:
            licencia_info['key'] = key_match.group(1)
        else:
            licencia_info['key'] = 'XXXXX'
    
    except Exception as e:
        print(f"Error detectando licencia Windows: {e}")
    
    return licencia_info


def detect_last_windows_update():
    """Detectar última actualización de Windows."""
    try:
        if not HAS_WINREG:
            return "No detectado"
        
        key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, 
                             r"SOFTWARE\Microsoft\Windows\CurrentVersion\WindowsUpdate\Auto Update\Results\Install")
        last_success = winreg.QueryValueEx(key, "LastSuccessTime")[0]
        winreg.CloseKey(key)
        
        # Formatear fecha
        if last_success:
            # Formato: YYYY-MM-DD HH:MM:SS
            try:
                date_obj = datetime.strptime(last_success, "%Y-%m-%d %H:%M:%S")
                return date_obj.strftime("%Y-%m-%d")
            except:
                return last_success[:10]  # Primeros 10 caracteres (fecha)
        
        return "No detectado"
    
    except Exception as e:
        return "No detectado"
    
def detect_mac_address():
    """Detectar dirección MAC de la interfaz de red principal."""
    try:
        import uuid
        mac = ':'.join(['{:02x}'.format((uuid.getnode() >> elements) & 0xff)
                       for elements in range(0,2*6,2)][::-1])
        return mac.upper()
    except:
        return "No detectado"


def detect_default_browser():
    """Detectar navegador predeterminado en Windows."""
    if not HAS_WINREG:
        return "No detectado"
    
    try:
        # Leer asociación de protocolo http
        key = winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\\Microsoft\\Windows\\Shell\\Associations\\UrlAssociations\\http\\UserChoice"
        )
        prog_id = winreg.QueryValueEx(key, "ProgId")[0]
        winreg.CloseKey(key)
        
        # Mapear ProgId a nombre de navegador
        browser_map = {
            'ChromeHTML': 'Google Chrome',
            'FirefoxURL': 'Mozilla Firefox',
            'MSEdgeHTM': 'Microsoft Edge',
            'IE.HTTP': 'Internet Explorer',
            'BraveHTML': 'Brave',
            'OperaStable': 'Opera'
        }
        
        for key_name, browser_name in browser_map.items():
            if key_name in prog_id:
                return browser_name
        
        return "Otro navegador"
    
    except Exception as e:
        # Fallback: buscar ejecutables comunes
        browsers = [
            (r"C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe", "Google Chrome"),
            (r"C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe", "Google Chrome"),
            (r"C:\\Program Files\\Mozilla Firefox\\firefox.exe", "Mozilla Firefox"),
            (r"C:\\Program Files (x86)\\Mozilla Firefox\\firefox.exe", "Mozilla Firefox"),
            (r"C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe", "Microsoft Edge"),
        ]
        
        for path, name in browsers:
            if os.path.exists(path):
                return f"{name} (detectado)"
        
        return "No detectado"


def detect_network_drives():
    """Detectar unidades de red mapeadas (ej: Z:\\, Y:\\)."""
    try:
        import subprocess
        
        # Ejecutar comando "net use" para listar unidades de red
        result = subprocess.run(
            ['net', 'use'],
            capture_output=True,
            text=True,
            timeout=5
        )
        
        output = result.stdout
        drives = []
        
        # Parsear salida de "net use"
        for line in output.split('\\n'):
            # Buscar líneas con unidades (formato: OK   Z:   \\\\servidor\\carpeta)
            if ':' in line and '\\\\\\\\' in line:
                parts = line.split()
                for part in parts:
                    if ':' in part and len(part) == 2:
                        drives.append(part)
        
        if drives:
            return ', '.join(sorted(set(drives)))
        else:
            return "Ninguna"
    
    except Exception as e:
        return "No detectado"


def detect_ip_local():
    """Detectar IP local del equipo."""
    try:
        # Método 1: Conectar a servidor externo (más confiable)
        import socket
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
        return local_ip
    except:
        try:
            # Método 2: Usar hostname
            hostname = socket.gethostname()
            local_ip = socket.gethostbyname(hostname)
            return local_ip
        except:
            return "No detectado"


# ============================================================================
# CLASE PRINCIPAL - INVENTORY MANAGER
# ============================================================================

class InventoryManagerApp:
    """Aplicación principal con CustomTkinter."""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Inventario Tecnológico - HRAJS")
        
        # Configurar tamaño de ventana inicial (1400x900 o 90% de pantalla)
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Usar 90% de pantalla o tamaño fijo (el menor)
        window_width = min(int(screen_width * 0.9), 1600)
        window_height = min(int(screen_height * 0.9), 1000)
        
        # Centrar ventana
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Variables de estado
        self.excel_path = None
        self.current_row = None
        self.current_sheet = "Equipos de Cómputo"  # Sheet actual
        self.equipment_data = {}
        self.verde_data = {}
        self.azul_data = {}
        
        # Widgets de formulario (para acceso posterior)
        self.manual_widgets = {}
        self.main_container = None  # Contenedor principal para cambiar vistas
        
        # PRIMERO: Crear menú nativo (por encima de todo)
        self.create_native_menu()
        
        # SEGUNDO: Crear header
        self.create_header()
        
        # TERCERO: Contenedor principal para las vistas
        self.main_container = ctk.CTkFrame(self.root, fg_color=COLOR_FONDO)
        self.main_container.pack(fill="both", expand=True, padx=0, pady=0)
        
        # CUARTO: Intentar cargar Excel automáticamente (después de que la ventana esté lista)
        self.root.after(100, self.auto_load_excel)
    
    
    def create_native_menu(self):
        """Crear menú nativo de tkinter."""
        
        # Crear barra de menú nativa
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # MENÚ ARCHIVO
        menu_archivo = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Archivo", menu=menu_archivo)
        menu_archivo.add_command(label="Cargar Excel", command=self.browse_excel)
        menu_archivo.add_separator()
        menu_archivo.add_command(label="Salir", command=self.root.quit)
        
        # MENÚ INVENTARIOS
        menu_inventarios = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Inventarios", menu=menu_inventarios)
        menu_inventarios.add_command(
            label="Equipos de Cómputo", 
            command=lambda: self.show_form_directo("Equipos de Cómputo")
        )
        menu_inventarios.add_command(
            label="Impresoras", 
            command=lambda: self.show_form_directo("Impresoras")
        )
        menu_inventarios.add_command(
            label="Periféricos", 
            command=lambda: self.show_form_directo("Periféricos")
        )
        menu_inventarios.add_command(
            label="Equipos de Red", 
            command=lambda: self.show_form_directo("Red")
        )
        
        # MENÚ OPERACIONES
        menu_operaciones = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Operaciones", menu=menu_operaciones)
        menu_operaciones.add_command(
            label="Mantenimiento", 
            command=lambda: self.show_form_directo("Mantenimiento")
        )
        menu_operaciones.add_command(
            label="Dar de Baja", 
            command=lambda: self.show_form_directo("Dados de Baja")
        )
        
        # MENÚ AYUDA
        menu_ayuda = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Ayuda", menu=menu_ayuda)
        menu_ayuda.add_command(label="Guía de Formulario", command=self.show_classification_guide)
    
    def show_form_directo(self, tipo):
        """Mostrar formulario directamente sin tabs."""
        if not self.excel_path:
            messagebox.showwarning("Advertencia", "Primero debes cargar un archivo Excel.\n\nVe a: Archivo → Cargar Excel")
            return
        
        # Limpiar contenedor principal
        for widget in self.main_container.winfo_children():
            widget.destroy()
        
        # Mostrar formulario correspondiente
        if tipo == "Equipos de Cómputo":
            self.show_manual_form_in_container()
        elif tipo == "Impresoras":
            self.create_impresoras_form_directo()
        elif tipo == "Periféricos":
            self.create_perifericos_form_directo()
        elif tipo == "Red":
            self.create_red_form_directo()
        elif tipo == "Mantenimiento":
            self.create_mantenimientos_form_directo()
        elif tipo == "Dados de Baja":
            self.create_baja_form_directo()
    
    def show_manual_form_in_container(self):
        """Mostrar formulario manual:"""
        # Limpiar contenedor
        for widget in self.main_container.winfo_children():
            widget.destroy()
        
        self.manual_widgets = {}
        
        # Frame scrollable
        form_frame = ctk.CTkScrollableFrame(
            self.main_container, 
            fg_color="#F5F5F5",
            corner_radius=0
        )
        form_frame.pack(fill="both", expand=True)
        
        # ===== TÍTULO =====
        title_frame = ctk.CTkFrame(form_frame, fg_color=COLOR_VERDE_HOSPITAL, corner_radius=12)
        title_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        # Obtener código siguiente
        try:
            next_code = self.get_next_codigo()
            codigo_text = f"Equipo: {next_code}"
        except:
            codigo_text = "Equipo: EQC-0001"
        
        title_label = ctk.CTkLabel(
            title_frame,
            text=f"{codigo_text}",
            font=("Segoe UI", 16, "bold"),
            text_color="white"
        )
        title_label.pack(pady=15)
        
        # ===== CAMPOS BÁSICOS =====
        self.create_form_field_centered(form_frame, "Tipo de Equipo", "tipo_equipo", 
                                        "combobox", TIPOS_EQUIPO)
        self.create_form_field_centered(form_frame, "Área / Servicio", "area_servicio", 
                                        "combobox", AREAS_SERVICIO)
        self.create_form_field_centered(form_frame, "Ubicación Específica", "ubicacion_especifica", 
                                        "entry")
        self.create_form_field_centered(form_frame, "Responsable / Custodio", "responsable_custodio", 
                                        "entry")
        
        # ===== SECCIÓN: CLASIFICACIÓN POR PROCESOS =====
        separator1 = ctk.CTkFrame(form_frame, height=2, fg_color="#CCCCCC")
        separator1.pack(fill="x", padx=20, pady=15)
        
        label_procesos = ctk.CTkLabel(
            form_frame,
            text="🏥 CLASIFICACIÓN POR PROCESOS",
            font=("Segoe UI", 14, "bold"),
            text_color=COLOR_AZUL_HOSPITAL
        )
        label_procesos.pack(pady=(10, 15))
        
        self.create_form_field_centered(form_frame, "Macroproceso", "macroproceso", 
                                        "combobox", list(MACROPROCESOS.keys()))
        self.create_form_field_centered(form_frame, "Proceso", "proceso", 
                                        "combobox", ["Selecciona primero Macroproceso"])
        self.create_form_field_centered(form_frame, "Subproceso", "subproceso", 
                                        "combobox", ["Selecciona primero Proceso"])
        
        # Configurar eventos condicionales
        self.manual_widgets['macroproceso'].configure(
            command=lambda choice: self.on_macroproceso_change(choice)
        )
        self.manual_widgets['proceso'].configure(
            command=lambda choice: self.on_proceso_change(choice)
        )
        
        # ===== SECCIÓN: SOFTWARE =====
        separator2 = ctk.CTkFrame(form_frame, height=2, fg_color="#CCCCCC")
        separator2.pack(fill="x", padx=20, pady=15)
        
        label_software = ctk.CTkLabel(
            form_frame,
            text="💻 SOFTWARE UTILIZADO",
            font=("Segoe UI", 14, "bold"),
            text_color=COLOR_AZUL_HOSPITAL
        )
        label_software.pack(pady=(10, 15))
        
        self.create_form_field_centered(form_frame, "Uso - SIHOS", "uso_sihos", 
                                        "combobox", SI_NO)
        self.create_form_field_centered(form_frame, "Uso - Office Básico", "uso_office_basico", 
                                        "combobox", SI_NO)
        self.create_form_field_centered(form_frame, "Software Especializado", "software_especializado", 
                                        "combobox", SI_NO)
        self.create_form_field_centered(form_frame, "Descripción Software", "descripcion_software", 
                                        "entry")
        self.create_form_field_centered(form_frame, "Función Principal", "funcion_principal", 
                                        "entry")
        
        # ===== SECCIÓN: INFORMACIÓN OPERATIVA =====
        separator4 = ctk.CTkFrame(form_frame, height=2, fg_color="#CCCCCC")
        separator4.pack(fill="x", padx=20, pady=15)
        
        label_operativo = ctk.CTkLabel(
            form_frame,
            text="⚙️ INFORMACIÓN OPERATIVA",
            font=("Segoe UI", 14, "bold"),
            text_color=COLOR_AZUL_HOSPITAL
        )
        label_operativo.pack(pady=(10, 15))
        
        self.create_form_field_centered(form_frame, "Horario de Uso", "horario_uso", 
                                        "combobox", HORARIOS_USO)
        self.create_form_field_centered(form_frame, "Estado Operativo", "estado_operativo", 
                                        "combobox", ESTADOS_OPERATIVOS)
        self.create_form_field_centered(form_frame, "Periodicidad Mantenimiento", "periodicidad_mtto", 
                                        "combobox", PERIODICIDADES_MTTO)
        self.create_form_field_centered(form_frame, "Responsable Mantenimiento", "responsable_mtto", 
                                        "combobox", TECNICOS_RESPONSABLES)
        self.create_form_field_centered(form_frame, "Observaciones Técnicas", "observaciones_tecnicas", 
                                        "entry")
        
        # ===== SECCIÓN: CUESTIONARIO CON RADIOBUTTONS =====
        from textos_tooltips import (
            CONF_LABELS, CONF_TOOLTIPS,
            INT_LABELS, INT_TOOLTIPS,
            CRIT_LABELS, CRIT_TOOLTIPS
        )
        
        separator3 = ctk.CTkFrame(form_frame, height=2, fg_color="#CCCCCC")
        separator3.pack(fill="x", padx=20, pady=15)
        
        label_cuestionario = ctk.CTkLabel(
            form_frame,
            text="🔒 CUESTIONARIO DE CLASIFICACIÓN (18 Preguntas Sí/No)",
            font=("Segoe UI", 14, "bold"),
            text_color=COLOR_AZUL_HOSPITAL
        )
        label_cuestionario.pack(pady=(10, 5))
        
        info_cuestionario = ctk.CTkLabel(
            form_frame,
            text="Selecciona Sí o No • Pasa el mouse sobre el texto para ver la pregunta completa",
            font=("Segoe UI", 11, "italic"),
            text_color="gray"
        )
        info_cuestionario.pack(pady=(0, 15))
        
        # CONFIDENCIALIDAD (9 preguntas - RadioButtons)
        label_conf = ctk.CTkLabel(
            form_frame,
            text="📋 CONFIDENCIALIDAD (Tipo de información):",
            font=("Segoe UI", 12, "bold"),
            text_color="#6F42C1"
        )
        label_conf.pack(anchor="w", padx=40, pady=(10, 5))
        
        for i, (label_text, tooltip_text) in enumerate(zip(CONF_LABELS, CONF_TOOLTIPS), 1):
            field_name = f"conf_{i}"
            self.create_radio_field_centered(form_frame, label_text, field_name, tooltip_text)
        
        # INTEGRIDAD (3 preguntas - RadioButtons)
        label_int = ctk.CTkLabel(
            form_frame,
            text="🔐 INTEGRIDAD (Compromiso de información):",
            font=("Segoe UI", 12, "bold"),
            text_color="#FD7E14"
        )
        label_int.pack(anchor="w", padx=40, pady=(15, 5))
        
        for i, (label_text, tooltip_text) in enumerate(zip(INT_LABELS, INT_TOOLTIPS), 1):
            field_name = f"int_{i}"
            self.create_radio_field_centered(form_frame, label_text, field_name, tooltip_text)
        
        # CRITICIDAD (6 preguntas - RadioButtons)
        label_crit = ctk.CTkLabel(
            form_frame,
            text="⚠️ CRITICIDAD (Impacto operacional):",
            font=("Segoe UI", 12, "bold"),
            text_color="#DC3545"
        )
        label_crit.pack(anchor="w", padx=40, pady=(15, 5))
        
        for i, (label_text, tooltip_text) in enumerate(zip(CRIT_LABELS, CRIT_TOOLTIPS), 1):
            field_name = f"crit_{i}"
            self.create_radio_field_centered(form_frame, label_text, field_name, tooltip_text)
        
        # ===== BOTONES (3 HORIZONTALES IGUALES) =====
        separator5 = ctk.CTkFrame(form_frame, height=2, fg_color="#CCCCCC")
        separator5.pack(fill="x", padx=20, pady=15)
        
        btn_action_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        btn_action_frame.pack(pady=20)
        
        BTN_WIDTH = 350
        BTN_HEIGHT = 50
        
        # Botón 1: GUARDAR
        self.btn_save_equipo = ctk.CTkButton(
            btn_action_frame,
            text="💾 GUARDAR",
            command=self.save_equipo_manual_only,
            font=("Segoe UI", 13, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5039",
            height=BTN_HEIGHT,
            width=BTN_WIDTH
        )
        self.btn_save_equipo.pack(side="left", padx=8)
        
        # Botón 2: ACTUALIZAR
        btn_update = ctk.CTkButton(
            btn_action_frame,
            text="🔄 ACTUALIZAR",
            command=self.update_equipo_computo,
            font=("Segoe UI", 13, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5039",
            height=BTN_HEIGHT,
            width=BTN_WIDTH
        )
        btn_update.pack(side="left", padx=8)
        
        # Botón 3: RECOPILACIÓN AUTOMÁTICA
        btn_collect = ctk.CTkButton(
            btn_action_frame,
            text="➡️ RECOPILACIÓN AUTO",
            command=self.start_automatic_collection,
            font=("Segoe UI", 13, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5039",
            height=BTN_HEIGHT,
            width=BTN_WIDTH
        )
        btn_collect.pack(side="left", padx=8)

    def on_macroproceso_change(self, selected_macroproceso):
        """Actualizar lista de Procesos cuando cambia el Macroproceso."""
        # Limpiar proceso y subproceso
        self.manual_widgets["proceso"].set("")
        self.manual_widgets["subproceso"].set("")
        self.manual_widgets["subproceso"].configure(values=[])
        
        # Obtener procesos del macroproceso seleccionado
        from config_listas import get_procesos_por_macroproceso
        procesos = get_procesos_por_macroproceso(selected_macroproceso)
        
        # Actualizar lista de procesos
        self.manual_widgets["proceso"].configure(values=procesos)


    def on_proceso_change(self, selected_proceso):
        """Actualizar lista de Subprocesos cuando cambia el Proceso."""
        # Limpiar subproceso
        self.manual_widgets["subproceso"].set("")
        
        # Obtener macroproceso actual
        macroproceso = self.manual_widgets["macroproceso"].get()
        
        if not macroproceso:
            return
        
        # Obtener subprocesos
        from config_listas import get_subprocesos_por_proceso
        subprocesos = get_subprocesos_por_proceso(macroproceso, selected_proceso)
        
        # Actualizar lista de subprocesos
        self.manual_widgets["subproceso"].configure(values=subprocesos)

    def create_impresoras_form_directo(self):
        """Crear formulario de impresoras directamente."""
        for widget in self.main_container.winfo_children():
            widget.destroy()
        self.create_impresoras_form(self.main_container)
    
    def create_perifericos_form_directo(self):
        """Crear formulario de periféricos directamente."""
        for widget in self.main_container.winfo_children():
            widget.destroy()
        self.create_perifericos_form(self.main_container)
    
    def create_red_form_directo(self):
        """Crear formulario de red directamente."""
        for widget in self.main_container.winfo_children():
            widget.destroy()
        self.create_red_form(self.main_container)
    
    def create_mantenimientos_form_directo(self):
        """Crear formulario de mantenimientos directamente."""
        for widget in self.main_container.winfo_children():
            widget.destroy()
        self.create_mantenimientos_form(self.main_container)
    
    def create_baja_form_directo(self):
        """Crear formulario de baja directamente."""
        for widget in self.main_container.winfo_children():
            widget.destroy()
        self.create_baja_form(self.main_container)

    def get_next_available_row(self, sheet_name, check_column=1, max_rows=500):
        """
        Función optimizada para buscar siguiente fila disponible en cualquier hoja.
        
        Args:
            sheet_name: Nombre de la hoja Excel
            check_column: Columna a verificar (default 1 = Consecutivo)
            max_rows: Máximo de filas a buscar (default 500)
        
        Returns:
            int: Número de la siguiente fila disponible
        """
        if not self.excel_path or not HAS_OPENPYXL:
            return 2
        
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            ws = wb[sheet_name]
            
            for row in range(2, max_rows + 2):
                if ws.cell(row=row, column=check_column).value is None:
                    wb.close()
                    return row
            
            wb.close()
            return max_rows + 2
            
        except Exception as e:
            print(f"Error buscando siguiente fila: {e}")
            return 2
    
    def create_header(self):
        """Crear encabezado con logo en círculo blanco."""
        header_frame = ctk.CTkFrame(self.root, fg_color=COLOR_VERDE_HOSPITAL, corner_radius=0)
        header_frame.pack(fill="x", padx=0, pady=0)
        
        # Frame interno
        content_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        content_frame.pack(pady=18)
        
        # Intentar cargar logo
        if HAS_PIL:
            logo_paths = [
                "logo_hospital.png",
                "logo.png", 
                "hospital_logo.png"
            ]
            
            for logo_path in logo_paths:
                if os.path.exists(logo_path):
                    try:
                        from PIL import Image, ImageDraw
                        
                        # 1. Cargar logo original
                        logo_original = Image.open(logo_path)
                        
                        # 2. Convertir a RGBA si no lo es
                        if logo_original.mode != 'RGBA':
                            logo_original = logo_original.convert('RGBA')
                        
                        # 3. Redimensionar logo
                        aspect_ratio = logo_original.width / logo_original.height
                        new_height = 65
                        new_width = int(new_height * aspect_ratio)
                        logo_resized = logo_original.resize((new_width, new_height), Image.Resampling.LANCZOS)
                        
                        # 4. Crear círculo blanco de fondo
                        circle_size = 90
                        background = Image.new('RGBA', (circle_size, circle_size), (0, 0, 0, 0))
                        
                        # Dibujar círculo blanco sólido
                        draw = ImageDraw.Draw(background)
                        draw.ellipse([0, 0, circle_size-1, circle_size-1], 
                                    fill=(255, 255, 255, 255))
                        
                        # 5. Centrar logo sobre círculo blanco
                        x_offset = (circle_size - new_width) // 2
                        y_offset = (circle_size - new_height) // 2
                        
                        # Pegar logo
                        background.paste(logo_resized, (x_offset, y_offset), logo_resized)
                        
                        # 6. Convertir a CTkImage
                        logo_ctk = ctk.CTkImage(
                            light_image=background, 
                            dark_image=background, 
                            size=(circle_size, circle_size)
                        )
                        
                        logo_label = ctk.CTkLabel(
                            content_frame,
                            image=logo_ctk,
                            text=""
                        )
                        logo_label.pack(side="left", padx=(0, 25))
                        print(f"✓ Logo cargado en círculo blanco: {logo_path}")
                        break
                        
                    except Exception as e:
                        print(f"✗ Error al cargar logo {logo_path}: {e}")
        
        # Texto del header
        text_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        text_frame.pack(side="left")
        
        title_label = ctk.CTkLabel(
            text_frame,
            text="SISTEMA DE INVENTARIO TECNOLÓGICO",
            font=("Segoe UI", 22, "bold"),
            text_color="white"
        )
        title_label.pack()
        
        subtitle_label = ctk.CTkLabel(
            text_frame,
            text="Hospital Regional Alfonso Jaramillo Salazar - Líbano, Tolima",
            font=("Segoe UI", 12),
            text_color="white"
        )
        subtitle_label.pack(pady=(2, 0))
        
        # Status label
        self.status_label = ctk.CTkLabel(
            header_frame,
            text="",
            font=("Segoe UI", 10),
            text_color="white",
            fg_color="transparent"
        )
        self.status_label.place(relx=0.98, rely=0.5, anchor="e")
    
    def browse_excel(self):
        """Abrir diálogo para seleccionar Excel."""
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo Excel - inventario_hospital_v1.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filename:
            self.excel_path = filename
            
            # Detectar siguiente fila
            self.current_row = self.get_next_available_row("Equipos de Cómputo", check_column=1)
            self.current_row = self.current_row-1
            
            # Actualizar status
            filename_short = os.path.basename(filename)
            self.status_label.configure(text=f"✅ {filename_short} cargado")
            
            # Mostrar pestañas
            self.show_manual_form_in_container()
            
            messagebox.showinfo("Éxito", f"✅ Archivo cargado correctamente:\n{filename_short}\n\nSiguiente fila disponible: {self.current_row}")
    
    def auto_load_excel(self):
        """Cargar Excel automáticamente si existe en el directorio actual."""
        default_file = "inventario_hospital_v1.xlsx"
        
        if os.path.exists(default_file):
            self.excel_path = default_file
            
            # Detectar siguiente fila automáticamente
            self.current_row = self.get_next_available_row("Equipos de Cómputo", check_column=1)
            self.current_row = self.current_row-1
            
            # Actualizar status
            self.status_label.configure(text=f"✅ {default_file} cargado")
            
            # Mostrar pestañas directamente
            self.show_manual_form_in_container()
            
            print(f"✅ Excel cargado automáticamente: {default_file}")
            print(f"✅ Siguiente fila disponible: {self.current_row}")
        else:
            # No hay archivo, mostrar mensaje en contenedor
            self.show_no_file_message()
    
    def show_no_file_message(self):
        """Mostrar mensaje cuando no hay archivo cargado."""
        for widget in self.main_container.winfo_children():
            widget.destroy()
        
        msg_frame = ctk.CTkFrame(self.main_container, fg_color=COLOR_FONDO)
        msg_frame.pack(fill="both", expand=True)
        
        # Centrar mensaje
        center_frame = ctk.CTkFrame(msg_frame, fg_color="transparent")
        center_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        icon_label = ctk.CTkLabel(
            center_frame,
            text="📂",
            font=("Segoe UI", 80)
        )
        icon_label.pack(pady=(0, 20))
        
        title_label = ctk.CTkLabel(
            center_frame,
            text="No se encontró el archivo Excel",
            font=("Segoe UI", 24, "bold"),
            text_color=COLOR_VERDE_HOSPITAL
        )
        title_label.pack(pady=(0, 10))
        
        subtitle_label = ctk.CTkLabel(
            center_frame,
            text="El sistema busca: inventario_hospital_v1.xlsx\nen el directorio actual",
            font=("Segoe UI", 14),
            text_color="#666666"
        )
        subtitle_label.pack(pady=(0, 30))
        
        btn_cargar = ctk.CTkButton(
            center_frame,
            text="📁 CARGAR ARCHIVO EXCEL",
            command=self.browse_excel,
            font=("Segoe UI", 16, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5A32",
            height=60,
            width=300,
            corner_radius=12
        )
        btn_cargar.pack()
    
    def detect_next_code(self, sheet_name, prefix):
        """Detectar siguiente código disponible basado en el último consecutivo en columna 1."""
        if not self.excel_path or not HAS_OPENPYXL:
            return f"{prefix}-001"
        
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            
            # Verificar que la hoja existe
            if sheet_name not in wb.sheetnames:
                wb.close()
                print(f"⚠️ Advertencia: Hoja '{sheet_name}' no existe. Creándola...")
                return f"{prefix}-001"
            
            ws = wb[sheet_name]
            
            # Buscar el ÚLTIMO consecutivo en columna 1 (no asumir que es next_row - 1)
            last_consecutive = 0
            for row in range(2, 500):
                value = ws.cell(row=row, column=1).value
                if value is not None:
                    try:
                        consecutivo = int(value)
                        if consecutivo > last_consecutive:
                            last_consecutive = consecutivo
                    except:
                        pass
                else:
                    break  # Primera fila vacía, detener
            
            wb.close()
            
            next_consecutive = last_consecutive + 1
            
            # Todos los códigos son de 4 dígitos
            return f"{prefix}-{next_consecutive:04d}"
            
        except Exception as e:
            print(f"❌ Error detectando código: {e}")
            import traceback
            traceback.print_exc()
            return f"{prefix}-001"
        
    def get_next_codigo(self):
        """
        Obtener siguiente código EQC para equipos de cómputo.
        Wrapper de detect_next_code() para compatibilidad.
        """
        return self.detect_next_code("Equipos de Cómputo", "EQC")
    
    def get_next_consecutivo(self):
        """
        Obtener siguiente número consecutivo para equipos de cómputo.
        Busca el último consecutivo en la columna 1 de "Equipos de Cómputo".
        """
        if not self.excel_path or not HAS_OPENPYXL:
            return 1
        
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            
            # Verificar que la hoja existe
            if "Equipos de Cómputo" not in wb.sheetnames:
                wb.close()
                return 1
            
            ws = wb["Equipos de Cómputo"]
            
            # Buscar el ÚLTIMO consecutivo en columna 1
            last_consecutive = 0
            for row in range(2, 500):
                value = ws.cell(row=row, column=1).value
                if value is not None:
                    try:
                        consecutivo = int(value)
                        if consecutivo > last_consecutive:
                            last_consecutive = consecutivo
                    except:
                        pass
                else:
                    break  # Primera fila vacía, detener
            
            wb.close()
            
            # Siguiente consecutivo
            return last_consecutive + 1
            
        except Exception as e:
            print(f"❌ Error obteniendo consecutivo: {e}")
            return 1
    
    def detect_next_consecutive_mantenimiento(self):
        """Detectar siguiente consecutivo para mantenimientos."""
        next_row = self.get_next_available_row("Mantenimientos", check_column=1)
        return next_row - 1
    
    def detect_next_baja(self):
        """Detectar siguiente número de baja."""
        next_row = self.get_next_available_row("Equipos Dados de Baja", check_column=1, max_rows=200)
        return next_row - 1
    
    def create_form_field_centered(self, parent, label_text, field_name, field_type, 
                               options=None, tooltip_text=None):
        """
        Crear campo centrado.
        
        Args:
            parent: Frame padre donde se creará el campo
            label_text: Texto del label (izquierda)
            field_name: Nombre del campo (key en self.manual_widgets)
            field_type: "entry" o "combobox"
            options: Lista de opciones para combobox (opcional)
            tooltip_text: Texto del tooltip al pasar mouse (opcional)
        
        Returns:
            widget: El widget creado (Entry o ComboBox)
        
        Grid:
            - Columna 0 (Label): 60% peso, alineado izquierda
            - Columna 1 (Widget): 40% peso, ancho fijo 300px
            - Espacio entre columnas: 30px
        """
        # Frame principal - CENTRADO con padding lateral
        field_frame = ctk.CTkFrame(parent, fg_color="white", corner_radius=8)
        field_frame.pack(fill="x", padx=40, pady=6)
        
        # Frame interno - Grid layout con espacio
        inner_frame = ctk.CTkFrame(field_frame, fg_color="transparent")
        inner_frame.pack(fill="x", padx=20, pady=10)
        
        # Configurar grid (2 columnas con espacio)
        inner_frame.grid_columnconfigure(0, weight=6)  # Label: flexible
        inner_frame.grid_columnconfigure(1, weight=0, minsize=300)  # Widget: fijo 300px
        
        # ===== LABEL (COLUMNA 0) =====
        label = ctk.CTkLabel(
            inner_frame,
            text=label_text,
            font=("Segoe UI", 12, "bold"),
            anchor="w",  # Alineado a la izquierda
            text_color="#333333"
        )
        label.grid(row=0, column=0, sticky="w", padx=(0, 30))  # 30px espacio a la derecha
        
        # Tooltip si existe
        if tooltip_text:
            ToolTip(label, tooltip_text)
        
        # ===== WIDGET (COLUMNA 1) =====
        if field_type == "combobox":
            widget = ctk.CTkComboBox(
                inner_frame,
                values=options if options else [],
                height=35,
                width=300, 
                font=("Segoe UI", 11),
                dropdown_font=("Segoe UI", 10),
                border_color="#CCCCCC",
                button_color=COLOR_VERDE_HOSPITAL,
                button_hover_color="#1F5A32",
                corner_radius=8
            )
            widget.grid(row=0, column=1, sticky="e")  # Alineado a la derecha de su columna
            
        elif field_type == "entry":
            widget = ctk.CTkEntry(
                inner_frame,
                height=35,
                width=300,
                font=("Segoe UI", 11),
                border_color="#CCCCCC",
                fg_color="white",
                corner_radius=8
            )
            widget.grid(row=0, column=1, sticky="e")  # Alineado a la derecha de su columna
        
        # Guardar widget
        self.manual_widgets[field_name] = widget
        return widget
    
    def create_date_field_centered(self, parent, label_text, field_name, tooltip_text=None):
        """
        Crear campo de FECHA con calendario (DateEntry).
        Similar a create_form_field_centered pero con calendario.
        """
        # Frame principal - CENTRADO con padding lateral
        field_frame = ctk.CTkFrame(parent, fg_color="white", corner_radius=8)
        field_frame.pack(fill="x", padx=40, pady=6)
        
        # Frame interno - Grid layout
        inner_frame = ctk.CTkFrame(field_frame, fg_color="transparent")
        inner_frame.pack(fill="x", padx=20, pady=10)
        
        # Configurar grid (2 columnas)
        inner_frame.grid_columnconfigure(0, weight=6)  # Label: flexible
        inner_frame.grid_columnconfigure(1, weight=0, minsize=300)  # Widget: fijo 300px
        
        # ===== LABEL (COLUMNA 0) =====
        label = ctk.CTkLabel(
            inner_frame,
            text=label_text,
            font=("Segoe UI", 12, "bold"),
            anchor="w",
            text_color="#333333"
        )
        label.grid(row=0, column=0, sticky="w", padx=(0, 30))
        
        # Tooltip si existe
        if tooltip_text:
            ToolTip(label, tooltip_text)
        
        # ===== WIDGET FECHA (COLUMNA 1) =====
        # Usar DateEntry (calendario visual)
        widget = DateEntry(
            inner_frame,
            width=28,
            background=COLOR_VERDE_HOSPITAL,
            foreground='white',
            borderwidth=2,
            font=("Segoe UI", 11),
            date_pattern='yyyy-mm-dd',  # Formato ISO
            showweeknumbers=False,
            showothermonthdays=False,
            selectbackground=COLOR_VERDE_HOSPITAL,
            selectforeground='white',
            normalbackground='white',
            normalforeground='black',
            weekendbackground='#F0F0F0',
            weekendforeground='black',
            othermonthbackground='white',
            othermonthweforeground='gray',
            othermonthwebackground='#F0F0F0'
        )
        widget.grid(row=0, column=1, sticky="e")
        
        # Guardar widget
        self.manual_widgets[field_name] = widget
        return widget
    
    def get_date_value(self, widget):
        """
        Obtener valor de fecha de un widget (DateEntry o Entry).
        
        Returns:
            str: Fecha en formato YYYY-MM-DD o cadena vacía
        """
        try:
            date_obj = widget.get_date()
            return date_obj.strftime('%Y-%m-%d')
        except:
            return ''
    
    def create_radio_field_centered(self, parent, label_text, field_name, tooltip_text=None):
        """
        Crear campo con RadioButtons (para preguntas Sí/No).
        """
        # Frame principal - CENTRADO
        field_frame = ctk.CTkFrame(parent, fg_color="white", corner_radius=8)
        field_frame.pack(fill="x", padx=40, pady=6)
        
        # Frame interno - Grid layout
        inner_frame = ctk.CTkFrame(field_frame, fg_color="transparent")
        inner_frame.pack(fill="x", padx=20, pady=10)
        
        # Configurar grid con columna fija para RadioButtons
        inner_frame.grid_columnconfigure(0, weight=6)  # Label: flexible
        inner_frame.grid_columnconfigure(1, weight=0, minsize=150)  # RadioButtons: fijo 150px
        
        # ===== LABEL (COLUMNA 0) =====
        label = ctk.CTkLabel(
            inner_frame,
            text=label_text,
            font=("Segoe UI", 12, "bold"),
            anchor="w",  # Alineado a la izquierda
            text_color="#333333"
        )
        label.grid(row=0, column=0, sticky="w", padx=(0, 30))  # 30px espacio a la derecha
        
        # Tooltip si existe
        if tooltip_text:
            ToolTip(label, tooltip_text)
        
        # ===== FRAME PARA RADIOBUTTONS (COLUMNA 1) =====
        radio_frame = ctk.CTkFrame(inner_frame, fg_color="transparent", width=150)
        radio_frame.grid(row=0, column=1, sticky="e")  # Alineado a la derecha
        radio_frame.grid_propagate(False)  # Mantener ancho fijo
        
        # Variable para almacenar selección
        var = tk.StringVar(value="")
        
        # RadioButton SÍ
        radio_si = ctk.CTkRadioButton(
            radio_frame,
            text="Sí",
            variable=var,
            value="Sí",
            font=("Segoe UI", 12),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5A32",
            border_width_checked=8,
            border_width_unchecked=2,
            width=60  # Ancho fijo para consistencia
        )
        radio_si.pack(side="left", padx=(0, 10))  # 10px entre Sí y No
        
        # RadioButton NO
        radio_no = ctk.CTkRadioButton(
            radio_frame,
            text="No",
            variable=var,
            value="No",
            font=("Segoe UI", 12),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5A32",
            border_width_checked=8,
            border_width_unchecked=2,
            width=60  # Ancho fijo para consistencia
        )
        radio_no.pack(side="left")
        
        # Guardar variable
        self.manual_widgets[field_name] = var
        
        return var
    
    def show_classification_guide(self):
        """Mostrar guía de formulario."""
        guide_window = ctk.CTkToplevel(self.root)
        guide_window.title("Guía del Formulario")
        guide_window.geometry("1000x700")
        
        # Centrar
        guide_window.update_idletasks()
        x = (guide_window.winfo_screenwidth() // 2) - 500
        y = (guide_window.winfo_screenheight() // 2) - 350
        guide_window.geometry(f"1000x700+{x}+{y}")
        
        # Header
        header_frame = ctk.CTkFrame(guide_window, fg_color=COLOR_VERDE_HOSPITAL, corner_radius=0)
        header_frame.pack(fill="x", padx=0, pady=0)
        
        header = ctk.CTkLabel(
            header_frame,
            text="📋 GUÍA DEL FORMULARIO",
            font=("Segoe UI", 24, "bold"),
            text_color="white"
        )
        header.pack(pady=(18, 5))
        
        subtitle = ctk.CTkLabel(
            header_frame,
            text="Sistema de Inventario Tecnológico v2.0",
            font=("Segoe UI", 12),
            text_color="white"
        )
        subtitle.pack(pady=(0, 18))
        
        # Tabview
        tabview = ctk.CTkTabview(guide_window, width=950, height=530)
        tabview.pack(pady=15, padx=25)
        
        # Tabs
        tabview.add("🔄 Flujo del Proceso")
        tabview.add("🏥 Macroproceso/Proceso")
        tabview.add("💡 Tips para Campos")
        
        # ===== TAB 1: FLUJO DEL PROCESO =====
        self._create_flujo_tab(tabview.tab("🔄 Flujo del Proceso"))
        
        # ===== TAB 2: MACROPROCESO/PROCESO =====
        self._create_macroproceso_tab(tabview.tab("🏥 Macroproceso/Proceso"))
        
        # ===== TAB 3: TIPS =====
        self._create_tips_tab(tabview.tab("💡 Tips para Campos"))
        
        # Botón cerrar
        btn_close = ctk.CTkButton(
            guide_window,
            text="✓ ENTENDIDO",
            command=guide_window.destroy,
            font=("Segoe UI", 14, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5A32",
            height=45,
            width=200
        )
        btn_close.pack(pady=15)

    def _create_flujo_tab(self, parent):
        """Tab de flujo del proceso."""
        scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        content = """
    🔄 FLUJO DEL PROCESO DE INVENTARIO

    1️⃣ DATOS MANUALES
    • Información básica del equipo
    • Área, ubicación, responsable
    • Software utilizado (SIFAX, Office, Especializado)
    • Información operativa (horarios, estado)
    • Macroproceso → Proceso → Subproceso (condicional)
    • Cuestionario de clasificación (18 preguntas)
    
    
    → Al terminar, puedes:
        💾 GUARDAR: Solo datos manuales (rápido)
        🔄 ACTUALIZAR: Modificar equipo existente
        ➡️ CONTINUAR: Pasar a detección automática

    2️⃣ RECOPILACIÓN AUTOMÁTICA
    • Detección de hardware (WMI)
        - Marca, Modelo, Serial del equipo, Discos
    • Sistema operativo y arquitectura
    • RAM y procesador
    • Software instalado (Office, Teams, Outlook)
    • Licencias de Windows y Office
    • Red (IP, conexión)
    • Seguridad (Antivirus, actualizaciones)
    
    ⚠️ Este proceso tarda 10-30 segundos

    3️⃣ VALIDACIÓN MIXTA
    • Revisar y corregir datos detectados
    • Agregar información de red (Switch, VLAN)
    • Configurar acceso remoto (AnyDesk)
    • Validar seguridad (Antivirus, Cifrado)

    4️⃣ GUARDADO FINAL
    • Se guardan TODAS las columnas en Excel
    • Archivo: inventario_hospital_v1.xlsx

    🎯 RECOMENDACIONES

    ✓ Usa GUARDAR si no necesitas detección automática
    ✓ Usa RECOPILACIÓN AUTO para inventario completo
    ✓ Puedes ACTUALIZAR equipos en cualquier momento
    ✓ Pasa el mouse sobre preguntas para ver texto completo
        """
        
        label = ctk.CTkLabel(
            scroll,
            text=content,
            font=("Consolas", 11),
            justify="left",
            anchor="w"
        )
        label.pack(fill="both", padx=20, pady=10)

    def _create_macroproceso_tab(self, parent):
        """Tab de explicación de Macroproceso."""
        scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        content = """
    🏥 MACROPROCESO → PROCESO → SUBPROCESO

    Esta estructura clasifica el equipo según el proceso hospitalario.
    Las listas son CONDICIONALES (se actualizan automáticamente).

    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    📊 ESTRATÉGICO (Dirección del hospital)
    → GERENCIA
        • Direccionamiento Estratégico
        • Asignación de Recursos
        • Evaluación y Desempeño
        • Rendición de Cuentas
        • Asesoría Jurídica
    
    → PLANEACIÓN Y CALIDAD
        • Seguridad del Paciente
        • Epidemiología
        • Estadística
        • Formulación de Planes
        • Revisión Documental

    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    🏥 MISIONAL (Atención de pacientes)
    → AMBULATORIA
        • Consulta Externa
        • Optometría
        • Odontología
        • Atención Hospitalizada
        • Atención Quirúrgica
    
    → SOPORTE DIAGNÓSTICO
        • Laboratorio Clínico
        • Imágenes Diagnósticas
        • Farmacia
        • Fisioterapia
        • Trabajo Social
    
    → URGENCIAS
        • Atención de Urgencias
        • Referencia y Contrarreferencia

    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    🔧 APOYO (Soporte operativo)
    → FINANCIERA
        • Facturación, Contabilidad, Cartera
    
    → TALENTO HUMANO
        • Recursos Humanos, Nómina
    
    → INFORMACIÓN Y COMUNICACIONES
        • Sistemas, SIAU, Archivo
    
    → AMBIENTE FÍSICO Y TECNOLOGÍA
        • Mantenimiento, Almacén, Servicios

    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    📋 EVALUACIÓN Y CONTROL
    → CONTROL INTERNO
    → AUDITORÍA MÉDICA
        """
        
        label = ctk.CTkLabel(
            scroll,
            text=content,
            font=("Consolas", 10),
            justify="left",
            anchor="w"
        )
        label.pack(fill="both", padx=20, pady=10)

    def _create_tips_tab(self, parent):
        """Tab de tips para campos."""
        scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        content = """
    💡 TIPS PARA LLENAR CAMPOS ESPECÍFICOS

    📅 FECHAS (Formato: YYYY-MM-DD)
    Ejemplo: 2024-01-15
    • Último Mantenimiento: Fecha del último mtto

    📋 CAMPOS OBLIGATORIOS (*)
    • Tipo de Equipo, Área, Ubicación
    • Responsable/Custodio
    • Macroproceso, Proceso, Subproceso
    • Uso SIHOS
    • Estado Operativo

    🔒 CUESTIONARIO (18 PREGUNTAS)
    • Responde Sí o No según corresponda
    • Pasa el mouse sobre cada pregunta para ver texto completo
    • Si tienes dudas, consulta con supervisor
    • Las respuestas se usan para clasificación posterior

    🏥 SOFTWARE ESPECIALIZADO
    • Marca "Sí" si usa programas especiales
    • En "Descripción", especifica cuáles:
        - PACS (Imágenes)
        - RIS (Radiología)
        - LIS (Laboratorio)
        - Contable específico
        - Otros sistemas propietarios

    ⏰ HORARIO DE USO
    • 24/7: Equipos que NUNCA se apagan (UCI, Urgencias)
    • Lun-Vie 7am-7pm: Atención extendida
    • Lun-Vie 7am-5pm: Horario administrativo

    🔧 MANTENIMIENTO
    • Periodicidad: Con qué frecuencia debe hacerse
    • Responsable: Técnico asignado
    • Último: Fecha del último realizado
    • Tipo: Preventivo, Correctivo o Predictivo

    ⚠️ ESTADO OPERATIVO
    • Operativo - Óptimo: Funciona perfecto
    • Operativo - Regular: Funciona con fallas menores
    • Operativo - Deficiente: Funciona mal, necesita atención
    • Fuera de Servicio: No funciona
    • En Reparación: En proceso de arreglo
        """
        
        label = ctk.CTkLabel(
            scroll,
            text=content,
            font=("Consolas", 10),
            justify="left",
            anchor="w"
        )
        label.pack(fill="both", padx=20, pady=10)        
    
    def start_automatic_collection(self):
        """Iniciar recopilación automática."""
        # Validar campos obligatorios
        required = ['tipo_equipo', 'area_servicio', 'ubicacion_especifica',
           'responsable_custodio', 'macroproceso', 'proceso', 'subproceso',
           'uso_sihos', 'estado_operativo']
        
        missing = []
        for field in required:
            widget = self.manual_widgets.get(field)
            if widget:
                try:
                    # Verificar que widget existe antes de acceder
                    if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                        value = widget.get().strip()
                        if not value or value == "Seleccionar...":
                            missing.append(field.replace('_', ' ').title())
                    else:
                        # Widget no existe, considerar campo faltante
                        missing.append(field.replace('_', ' ').title())
                except:
                    # Error al acceder al widget, considerar campo faltante
                    missing.append(field.replace('_', ' ').title())
        
        if missing:
            messagebox.showwarning(
                "Campos Incompletos",
                f"Debes completar los siguientes campos:\n\n" + "\n".join(f"• {m}" for m in missing)
            )
            return
        
        # Guardar datos manuales con verificación
        self.equipment_data = {}
        for field_name, widget in self.manual_widgets.items():
            try:
                if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                    value = widget.get().strip()
                    if value and value != "Seleccionar...":
                        self.equipment_data[field_name] = value
            except:
                pass  # Si falla, simplemente no guarda ese campo
        
        # Mostrar ventana de progreso
        self.show_progress_window()
        
        # Ejecutar recopilación en thread separado
        thread = threading.Thread(target=self.collect_automatic_data)
        thread.daemon = True
        thread.start()
    
    def show_progress_window(self):
        """Mostrar ventana de progreso."""
        self.progress_window = ctk.CTkToplevel(self.root)
        self.progress_window.title("Recopilación Automática")
        self.progress_window.geometry("600x400")
        self.progress_window.transient(self.root)
        self.progress_window.grab_set()
        
        # Centrar
        self.progress_window.update_idletasks()
        x = (self.progress_window.winfo_screenwidth() // 2) - 300
        y = (self.progress_window.winfo_screenheight() // 2) - 200
        self.progress_window.geometry(f"600x400+{x}+{y}")
        
        label = ctk.CTkLabel(
            self.progress_window,
            text="🔄 Recopilando Datos Automáticos...",
            font=("Arial", 16, "bold")
        )
        label.pack(pady=20)
        
        self.progress_bar = ctk.CTkProgressBar(
            self.progress_window,
            mode="indeterminate",
            width=500
        )
        self.progress_bar.pack(pady=10)
        self.progress_bar.start()
        
        self.log_text = ctk.CTkTextbox(
            self.progress_window,
            width=550,
            height=250,
            font=("Consolas", 10)
        )
        self.log_text.pack(pady=10, padx=20)
    
    def log_progress(self, message):
        """Agregar mensaje al log."""
        if hasattr(self, 'log_text'):
            self.log_text.insert("end", message + "\n")
            self.log_text.see("end")
            self.root.update()
    
    def collect_automatic_data(self):
        """Recopilar datos automáticos."""
        self.verde_data = {}
        
        # 1. Nombre del equipo
        self.log_progress("📋 Identificación del equipo...")
        self.verde_data['nombre_equipo'] = socket.gethostname()
        self.log_progress(f"   ✓ Nombre: {self.verde_data['nombre_equipo']}")
        
        # 2. Hardware con WMI
        self.log_progress("\n💻 Detectando hardware con WMI...")
        hw_info = detect_hardware_wmi()
        
        # Equipo
        self.verde_data['marca'] = hw_info['marca']
        self.verde_data['modelo'] = hw_info['modelo']
        self.verde_data['serial'] = hw_info['serial']
        
        self.log_progress(f"   ✓ Marca: {self.verde_data['marca']}")
        self.log_progress(f"   ✓ Modelo: {self.verde_data['modelo']}")
        self.log_progress(f"   ✓ Serial: {self.verde_data['serial']}")
        
        # ===== DISCO 1 (PRIMARIO) =====
        self.log_progress(f"\n💿 Disco 1 (Primario):")
        self.verde_data['disco1_capacidad'] = hw_info['disco1_capacidad']
        self.verde_data['disco1_tipo'] = hw_info['disco1_tipo']
        self.verde_data['disco1_serial'] = hw_info['disco1_serial']
        self.verde_data['disco1_marca'] = hw_info['disco1_marca']
        self.verde_data['disco1_modelo'] = hw_info['disco1_modelo']
        
        self.log_progress(f"   ✓ Capacidad: {hw_info['disco1_capacidad']} GB")
        self.log_progress(f"   ✓ Tipo: {hw_info['disco1_tipo']}")
        self.log_progress(f"   ✓ Serial: {hw_info['disco1_serial']}")
        self.log_progress(f"   ✓ Marca: {hw_info['disco1_marca']}")
        self.log_progress(f"   ✓ Modelo: {hw_info['disco1_modelo']}")
        
        # ===== DISCO 2 (SECUNDARIO) =====
        if hw_info['disco2_capacidad'] != 'No tiene':
            self.log_progress(f"\n💿 Disco 2 (Secundario) Detectado:")
            self.log_progress(f"   ✓ Capacidad: {hw_info['disco2_capacidad']} GB")
            self.log_progress(f"   ✓ Tipo: {hw_info['disco2_tipo']}")
            self.log_progress(f"   ✓ Serial: {hw_info['disco2_serial']}")
            self.log_progress(f"   ✓ Marca: {hw_info['disco2_marca']}")
            self.log_progress(f"   ✓ Modelo: {hw_info['disco2_modelo']}")
        else:
            self.log_progress(f"\n💿 Disco 2 (Secundario): No detectado")
        
        # 5-7. Sistema Operativo
        self.log_progress("\n🪟 Sistema Operativo...")
        self.verde_data['sistema_operativo'] = f"{platform.system()} {platform.release()}"
        self.verde_data['arquitectura_so'] = "64 bits" if "64" in platform.machine() else "32 bits"
        self.verde_data['procesador'] = platform.processor() or "No detectado"
        
        self.log_progress(f"   ✓ SO: {self.verde_data['sistema_operativo']}")
        self.log_progress(f"   ✓ Arquitectura: {self.verde_data['arquitectura_so']}")
        self.log_progress(f"   ✓ Procesador: {self.verde_data['procesador'][:50]}...")
        
        # 8-9. RAM y Almacenamiento
        if HAS_PSUTIL:
            # Obtener RAM utilizable
            ram_bytes = psutil.virtual_memory().total
            ram_gib_usable = ram_bytes / (1024**3)  # GiB utilizables
            
            # Tamaños comerciales estándar
            common_sizes = [2, 4, 6, 8, 12, 16, 24, 32, 48, 64, 128]
            
            # Redondeo inteligente con margen del 15%
            # Busca el tamaño comercial más probable
            # considerando que puede haber RAM reservada
            best_match = None
            min_diff = float('inf')
            
            for size in common_sizes:
                # Considerar margen de -20% (por GPU integrada, BIOS, etc)
                expected_usable = size * 0.80  # 80% del tamaño comercial
                diff = abs(ram_gib_usable - expected_usable)
                
                # También considerar coincidencia directa
                direct_diff = abs(ram_gib_usable - size)
                
                # Usar la mejor coincidencia
                actual_diff = min(diff, direct_diff)
                
                if actual_diff < min_diff:
                    min_diff = actual_diff
                    best_match = size
            
            ram_gb = best_match
            
            self.verde_data['ram_gb'] = str(ram_gb)
            self.log_progress(f"   ✓ RAM: {ram_gb} GB (utilizable: {ram_gib_usable:.2f} GiB)")
            
            # Almacenamiento
            try:
                disk = psutil.disk_usage('C:\\\\')
                storage_gb = round(disk.total / (1024**3))
                self.verde_data['almacenamiento_gb'] = str(storage_gb)
                self.log_progress(f"   ✓ Almacenamiento: {storage_gb} GB")
            except:
                self.verde_data['almacenamiento_gb'] = "No detectado"
        else:
            self.verde_data['ram_gb'] = "Requiere psutil"
            self.verde_data['almacenamiento_gb'] = "Requiere psutil"
        
        # 10-15. Software Office
        self.log_progress("\n📦 Detectando Office...")
        office_version, office_licencia = detect_office_version()
        self.verde_data['version_office'] = office_version
        self.verde_data['licencia_office'] = office_licencia
        self.verde_data['uso_navegador_web'] = "Sí"
        
        self.log_progress(f"   ✓ Versión Office: {office_version}")
        self.log_progress(f"   ✓ Licencia Office: {office_licencia}")
        
        # Teams y Outlook
        teams, outlook = detect_office_apps()
        self.verde_data['uso_teams'] = teams
        self.verde_data['uso_outlook'] = outlook
        
        self.log_progress(f"   ✓ Teams: {teams}")
        self.log_progress(f"   ✓ Outlook: {outlook}")
        
        # 16-18. Licencia Windows
        self.log_progress("\n🔑 Detectando licencia Windows...")
        lic_info = detect_windows_license()
        self.verde_data['licencia_windows'] = lic_info['tipo']
        self.verde_data['key_windows'] = lic_info['key']
        self.verde_data['estado_licencia_windows'] = lic_info['estado']
        
        self.log_progress(f"   ✓ Licencia: {lic_info['tipo']}")
        self.log_progress(f"   ✓ Key (últimos 5): {lic_info['key']}")
        self.log_progress(f"   ✓ Estado: {lic_info['estado']}")
        
        # 19-23. Red
        self.log_progress("\\n🌐 Red...")

        # IP Local
        ip_local = detect_ip_local()
        self.verde_data['direccion_ip'] = ip_local
        self.log_progress(f"   ✓ IP Local: {ip_local}")

        # MAC Address
        mac_address = detect_mac_address()
        self.verde_data['mac_address'] = mac_address
        self.log_progress(f"   ✓ MAC Address: {mac_address}")

        self.verde_data['tipo_conexion'] = "Ethernet"  # Default

        # Navegador predeterminado
        navegador = detect_default_browser()
        self.verde_data['navegador_predeterminado'] = navegador
        self.log_progress(f"   ✓ Navegador: {navegador}")

        # Unidades de red mapeadas (nuevo)
        unidades_red = detect_network_drives()
        self.verde_data['unidades_red_mapeadas'] = unidades_red
        self.log_progress(f"   ✓ Unidades de red: {unidades_red}")
        
        # 24-26. Seguridad
        self.log_progress("\n🔒 Seguridad...")
        self.verde_data['antivirus_instalado'] = "Windows Defender"
        self.verde_data['windows_update_activo'] = "Sí"
        
        last_update = detect_last_windows_update()
        self.verde_data['ultima_act_windows'] = last_update
        
        self.log_progress(f"   ✓ Antivirus: Windows Defender")
        self.log_progress(f"   ✓ Última actualización: {last_update}")
        
        self.log_progress("\n✅ Recopilación automática completada")
        
        # Cerrar ventana de progreso
        self.progress_bar.stop()
        self.root.after(1000, lambda: self.progress_window.destroy())
        
        # Mostrar validación de campos mixtos
        self.root.after(1500, self.show_mixed_validation)
    
    def show_mixed_validation(self):
        """Mostrar ventana de validación de campos mixtos (AZULES)."""
        validation_window = ctk.CTkToplevel(self.root)
        validation_window.title("Validación de Campos Mixtos")
        validation_window.geometry("900x600")
        validation_window.transient(self.root)
        validation_window.grab_set()
        
        # Centrar
        validation_window.update_idletasks()
        x = (validation_window.winfo_screenwidth() // 2) - 450
        y = (validation_window.winfo_screenheight() // 2) - 300
        validation_window.geometry(f"900x600+{x}+{y}")
        
        header = ctk.CTkLabel(
            validation_window,
            text="🔵 VALIDACIÓN DE CAMPOS MIXTOS (AZULES)",
            font=("Arial", 18, "bold"),
            text_color=COLOR_AZUL_HOSPITAL
        )
        header.pack(pady=20)
        
        info = ctk.CTkLabel(
            validation_window,
            text="Valida o corrige los siguientes campos:",
            font=("Arial", 13)
        )
        info.pack(pady=(0, 20))
        
        # Frame scrollable
        scroll_frame = ctk.CTkScrollableFrame(validation_window, width=850, height=380)
        scroll_frame.pack(pady=10, padx=25)
        
        # Campos mixtos (SIN DISCO 2)
        self.mixed_widgets = {}
        
        mixed_fields = [
            # RED Y ACCESO REMOTO
            ("Switch / Puerto", "switch_puerto", "entry", "No detectado"),
            ("VLAN Asignada", "vlan_asignada", "entry", "No detectado"),
            ("ID AnyDesk", "id_anydesk", "entry", self.detect_anydesk()),
            ("Otro Acceso Remoto", "otro_acceso_remoto", "entry", "Ninguno"),
            
            # SEGURIDAD
            ("Estado Antivirus", "estado_antivirus", "combobox", OPCIONES_ESTADO_ANTIVIRUS),
            ("Cifrado de Disco", "cifrado_disco", "combobox", OPCIONES_CIFRADO_DISCO),
            ("Tipo Usuario Local", "tipo_usuario_local", "combobox", OPCIONES_TIPO_USUARIO),
        ]
        
        for label_text, field_name, field_type, default in mixed_fields:
            field_frame = ctk.CTkFrame(scroll_frame, fg_color="transparent")
            field_frame.pack(fill="x", padx=15, pady=10)
            
            label = ctk.CTkLabel(
                field_frame,
                text=label_text,
                font=("Arial", 13, "bold"),
                width=250,
                anchor="w"
            )
            label.pack(side="left", padx=(0, 15))
            
            if field_type == "combobox":
                widget = ctk.CTkComboBox(
                    field_frame,
                    values=default if isinstance(default, list) else ["No detectado"],
                    width=500,
                    font=("Arial", 12),
                    dropdown_font=("Arial", 11),
                    height=32
                )
                if isinstance(default, list) and len(default) > 0:
                    widget.set(default[0])
            else:
                widget = ctk.CTkEntry(
                    field_frame,
                    width=500,
                    font=("Arial", 12),
                    height=32
                )
                widget.insert(0, str(default))
            
            widget.pack(side="left", fill="x", expand=True)
            self.mixed_widgets[field_name] = widget
        
        # Botón continuar
        btn_save = ctk.CTkButton(
            validation_window,
            text="✅ VALIDAR Y GUARDAR EN EXCEL",
            command=lambda: self.save_mixed_and_excel(validation_window),
            font=("Arial", 14, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5A32",
            height=45
        )
        btn_save.pack(pady=20, padx=20, fill="x")
    
    def detect_anydesk(self):
        """Detectar ID de AnyDesk si está instalado."""
        try:
            # Ruta típica de AnyDesk
            anydesk_path = r"C:\Program Files (x86)\AnyDesk\AnyDesk.exe"
            if os.path.exists(anydesk_path):
                # Intentar obtener ID (simplificado)
                return "Instalado - Verificar ID"
            return "No instalado"
        except:
            return "No detectado"
    
    def save_mixed_and_excel(self, validation_window):
        """Guardar datos mixtos y todo en Excel."""
        # Obtener datos de campos mixtos
        self.azul_data = {}
        for field_name, widget in self.mixed_widgets.items():
            value = widget.get().strip()
            if value:
                self.azul_data[field_name] = value
        
        # Cerrar ventana de validación
        validation_window.destroy()
        
        # Guardar en Excel
        self.save_to_excel()
        
        # Mostrar mensaje de completado
        self.show_completion_message()
    
    def save_to_excel(self):
        """Guardar TODOS los datos en Excel."""
        if not HAS_OPENPYXL:
            messagebox.showerror("Error", "Necesitas instalar openpyxl")
            return
        
        try:
            wb = load_workbook(self.excel_path)
            ws = wb["Equipos de Cómputo"]
            
            # Verificar modo
            if hasattr(self, 'equipo_update_row') and self.equipo_update_row:
                # MODO ACTUALIZACIÓN
                row = self.equipo_update_row
                codigo = self.equipo_update_code
                consecutive = int(codigo.split('-')[1])
            else:
                # MODO GUARDAR NUEVO
                row = self.current_row
                consecutive = row - 1
                ws.cell(row=row, column=1, value=consecutive)
                ws.cell(row=row, column=2, value=f"EQC-{consecutive:04d}")
            
            # ===== COLUMNA 3: Nombre Equipo (VERDE) =====
            ws.cell(row=row, column=3, value=self.verde_data.get('nombre_equipo', ''))
            
            # ===== COLUMNAS 4-50: NARANJAS (DATOS MANUALES) =====
            col = 4
            
            # Campos básicos (7)
            basic_fields = [
                'tipo_equipo', 'area_servicio', 'ubicacion_especifica', 'responsable_custodio',
                'macroproceso', 'proceso', 'subproceso'
            ]
            
            for field in basic_fields:
                value = self.equipment_data.get(field, '')
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            # Campos software (5)
            software_fields = ['uso_sihos', 'uso_office_basico', 'software_especializado', 
                            'descripcion_software', 'funcion_principal']
            for field in software_fields:
                value = self.equipment_data.get(field, '')
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            # ===== CUESTIONARIO DE CLASIFICACIÓN (18 PREGUNTAS) =====
            # 9 Confidencialidad
            for i in range(1, 10):
                value = self.equipment_data.get(f'conf_{i}', '')
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            # 3 Integridad
            for i in range(1, 4):
                value = self.equipment_data.get(f'int_{i}', '')
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            # 6 Criticidad
            for i in range(1, 7):
                value = self.equipment_data.get(f'crit_{i}', '')
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            # Campos finales (4) - SIN FECHAS NI VALORES
            final_fields = [
                'horario_uso', 'estado_operativo', 'observaciones_tecnicas',
                'periodicidad_mtto', 'responsable_mtto'
            ]
            for field in final_fields:
                value = self.equipment_data.get(field, '')
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            # TOTAL NARANJAS: 7 + 5 + 1 + 18 + 4 = 35 columnas
            # Básicos (7) + Software (5) + Función (1) + Cuestionario (18) + Operativos (4)
            
            # ===== COLUMNAS VERDES (HARDWARE Y SOFTWARE) =====
            verde_fields = [
            'marca', 'modelo', 'serial', 'sistema_operativo', 'arquitectura_so',
            'procesador', 'ram_gb',
            # DISCO 1 (5 campos)
            'disco1_capacidad', 'disco1_tipo', 'disco1_serial', 'disco1_marca', 'disco1_modelo',
            # DISCO 2 (5 campos)
            'disco2_capacidad', 'disco2_tipo', 'disco2_serial', 'disco2_marca', 'disco2_modelo',
            # Resto
            'uso_navegador_web', 'version_office', 'licencia_office',
            'uso_teams', 'uso_outlook', 'licencia_windows', 'key_windows',
            'estado_licencia_windows', 
            'direccion_ip',           # IP local
            'mac_address',           
            'tipo_conexion',
            'navegador_predeterminado',  
            'unidades_red_mapeadas',     
            'antivirus_instalado', 'ultima_act_windows', 'windows_update_activo'
        ]
            
            for field in verde_fields:
                value = self.verde_data.get(field, '')
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            # ===== COLUMNAS AZULES (MIXTAS) =====
            azul_fields = [
                'switch_puerto', 'vlan_asignada', 'id_anydesk',
                'otro_acceso_remoto', 'estado_antivirus',
                'cifrado_disco', 'tipo_usuario_local'
            ]
            
            for field in azul_fields:
                value = self.azul_data.get(field, '')
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            # Guardar
            wb.save(self.excel_path)
            wb.close()
            
            # Mensaje según modo
            if hasattr(self, 'equipo_update_row') and self.equipo_update_row:
                messagebox.showinfo("Éxito", f"✅ Equipo {codigo} actualizado correctamente (datos completos)")
                self.reset_after_update_equipos()
            else:
                messagebox.showinfo("Éxito", f"✅ Equipo guardado: EQC-{consecutive:04d}")
                self.current_row += 1
                self.root.after(100, self.show_manual_form_in_container)
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar en Excel:\n{e}")
    
    def save_equipo_manual_only(self):
        """Guardar solo datos manuales (sin recopilación automática).        """
        # ===== VALIDACIÓN DE CAMPOS OBLIGATORIOS =====
        required_fields = {
            'tipo_equipo': 'Tipo de Equipo',
            'area_servicio': 'Área / Servicio',
            'ubicacion_especifica': 'Ubicación Específica',
            'responsable_custodio': 'Responsable / Custodio',
            'macroproceso': 'Macroproceso',
            'proceso': 'Proceso',
            'subproceso': 'Subproceso',
            'uso_sihos': 'Uso SIHOS',
            'estado_operativo': 'Estado Operativo'
        }
        
        missing_fields = []
        
        for field_name, field_label in required_fields.items():
            if field_name in self.manual_widgets:
                widget = self.manual_widgets[field_name]
                try:
                    if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                        if isinstance(widget, (ctk.CTkEntry, ctk.CTkComboBox)):
                            value = widget.get().strip()
                        elif isinstance(widget, tk.StringVar):
                            value = widget.get().strip()
                        else:
                            value = ''
                        
                        if not value:
                            missing_fields.append(field_label)
                    else:
                        missing_fields.append(field_label)
                except:
                    missing_fields.append(field_label)
            else:
                missing_fields.append(field_label)
        
        if missing_fields:
            messagebox.showwarning(
                "Campos Requeridos",
                f"Por favor completa los siguientes campos obligatorios:\n\n" + 
                "\n".join(f"• {field}" for field in missing_fields)
            )
            return
        
        # ===== RECOPILAR DATOS MANUALES =====
        datos_guardados = {}
        
        for field_name, widget in self.manual_widgets.items():
            try:
                if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                    # Entry (campos de texto)
                    if isinstance(widget, ctk.CTkEntry):
                        datos_guardados[field_name] = widget.get()
                    
                    # ComboBox (listas desplegables)
                    elif isinstance(widget, ctk.CTkComboBox):
                        datos_guardados[field_name] = widget.get()
                    
                    # StringVar (RadioButtons) ← NUEVO
                    elif isinstance(widget, tk.StringVar):
                        datos_guardados[field_name] = widget.get()
                    
                    else:
                        datos_guardados[field_name] = ''
                else:
                    datos_guardados[field_name] = ''
            except Exception as e:
                print(f"Error al obtener valor de {field_name}: {e}")
                datos_guardados[field_name] = ''
        
        # ===== GUARDAR EN EXCEL =====
        try:
            # Abrir Excel
            if not os.path.exists(self.excel_path):
                messagebox.showerror("Error", f"No se encontró el archivo: {self.excel_path}")
                return
            
            wb = openpyxl.load_workbook(self.excel_path)
            
            if "Equipos de Cómputo" not in wb.sheetnames:
                messagebox.showerror("Error", f"No se encontró la hoja: {"Equipos de Cómputo"}")
                return
            
            ws = wb["Equipos de Cómputo"]
            
            # Obtener siguiente consecutivo y código
            next_consecutivo = self.get_next_consecutivo()
            next_codigo = self.get_next_codigo()
            
            # Nueva fila
            nueva_fila = ws.max_row + 1
            
            # ===== MAPEO A 85 COLUMNAS =====
            
            # Cols 1-2: Identificación
            ws.cell(row=nueva_fila, column=1).value = next_consecutivo  # N° Consecutivo
            ws.cell(row=nueva_fila, column=2).value = next_codigo       # Código
            
            # Col 3: Nombre Equipo (VERDE - se llenará después)
            ws.cell(row=nueva_fila, column=3).value = ''  # Vacío por ahora
            
            # Cols 4-7: Básicos (NARANJA)
            ws.cell(row=nueva_fila, column=4).value = datos_guardados.get('tipo_equipo', '')
            ws.cell(row=nueva_fila, column=5).value = datos_guardados.get('area_servicio', '')
            ws.cell(row=nueva_fila, column=6).value = datos_guardados.get('ubicacion_especifica', '')
            ws.cell(row=nueva_fila, column=7).value = datos_guardados.get('responsable_custodio', '')
            
            # Cols 8-10: Macroproceso/Proceso/Subproceso (NARANJA)
            ws.cell(row=nueva_fila, column=8).value = datos_guardados.get('macroproceso', '')
            ws.cell(row=nueva_fila, column=9).value = datos_guardados.get('proceso', '')
            ws.cell(row=nueva_fila, column=10).value = datos_guardados.get('subproceso', '')
            
            # Cols 11-16: Software (NARANJA)
            ws.cell(row=nueva_fila, column=11).value = datos_guardados.get('uso_sihos', '')
            ws.cell(row=nueva_fila, column=13).value = datos_guardados.get('uso_office_basico', '')
            ws.cell(row=nueva_fila, column=14).value = datos_guardados.get('software_especializado', '')
            ws.cell(row=nueva_fila, column=15).value = datos_guardados.get('descripcion_software', '')
            ws.cell(row=nueva_fila, column=16).value = datos_guardados.get('funcion_principal', '')
            
            # Cols 17-34: Cuestionario 18 preguntas (NARANJA)
            # CONFIDENCIALIDAD (9)
            for i in range(1, 10):
                ws.cell(row=nueva_fila, column=16 + i).value = datos_guardados.get(f'conf_{i}', '')
            
            # INTEGRIDAD (3)
            for i in range(1, 4):
                ws.cell(row=nueva_fila, column=25 + i).value = datos_guardados.get(f'int_{i}', '')
            
            # CRITICIDAD (6)
            for i in range(1, 7):
                ws.cell(row=nueva_fila, column=28 + i).value = datos_guardados.get(f'crit_{i}', '')
            
            # Cols 35-46: Operativos (NARANJA)
            ws.cell(row=nueva_fila, column=35).value = datos_guardados.get('horario_uso', '')
            ws.cell(row=nueva_fila, column=36).value = datos_guardados.get('estado_operativo', '')
            ws.cell(row=nueva_fila, column=37).value = datos_guardados.get('observaciones_tecnicas', '')
            ws.cell(row=nueva_fila, column=38).value = datos_guardados.get('periodicidad_mtto', '')
            ws.cell(row=nueva_fila, column=39).value = datos_guardados.get('responsable_mtto', '')
            
            # Cols 47-77: Hardware/Software (VERDE) - Vacíos por ahora
            for col in range(47, 78):
                ws.cell(row=nueva_fila, column=col).value = ''
            
            # Cols 78-84: Mixtos (AZUL) - Vacíos por ahora
            for col in range(78, 85):
                ws.cell(row=nueva_fila, column=col).value = ''
            
            # Guardar
            wb.save(self.excel_path)
            wb.close()
            
            messagebox.showinfo(
                "Éxito",
                f"Equipo guardado exitosamente:\n\n" +
                f"Código: {next_codigo}\n" +
                f"Tipo: {datos_guardados.get('tipo_equipo', '')}\n" +
                f"Área: {datos_guardados.get('area_servicio', '')}\n\n" +
                f"Los datos de hardware se pueden agregar después con 'Recopilación Automática'."
            )
            
            # Limpiar formulario
            for widget in self.manual_widgets.values():
                try:
                    if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                        if isinstance(widget, ctk.CTkEntry):
                            widget.delete(0, 'end')
                        elif isinstance(widget, ctk.CTkComboBox):
                            widget.set('')
                        elif isinstance(widget, tk.StringVar):
                            widget.set('')
                except:
                    pass
            
        except Exception as e:
            messagebox.showerror(
                "Error al Guardar",
                f"Ocurrió un error al guardar los datos:\n\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
    
    def update_equipo_computo(self):
        """Actualizar equipo de cómputo existente."""
        if not self.excel_path:
            messagebox.showerror("Error", "No hay Excel cargado")
            return
        
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Actualizar Equipo de Cómputo")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 200
        y = (dialog.winfo_screenheight() // 2) - 100
        dialog.geometry(f"400x200+{x}+{y}")
        
        ctk.CTkLabel(
            dialog,
            text="Ingresa el código del equipo a actualizar:",
            font=("Segoe UI", 13)
        ).pack(pady=20)
        
        entry_codigo = ctk.CTkEntry(
            dialog,
            width=200,
            height=40,
            font=("Segoe UI", 12),
            placeholder_text="Ej: EQC-0142"
        )
        entry_codigo.pack(pady=10)
        entry_codigo.focus()
        
        def buscar_y_cargar():
            codigo = entry_codigo.get().strip().upper()
            if not codigo:
                messagebox.showerror("Error", "Debes ingresar un código")
                return
            
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["Equipos de Cómputo"]
                
                found = False
                target_row = None
                
                for row in range(2, 500):
                    cell_value = ws.cell(row=row, column=2).value
                    if cell_value and cell_value.upper() == codigo:
                        found = True
                        target_row = row
                        break
                
                if not found:
                    wb.close()
                    messagebox.showerror("Error", f"No se encontró el código {codigo}")
                    return
                
                # Cargar datos NARANJAS (columnas 4-27)
                naranja_fields = [
                    'tipo_equipo', 'area_servicio', 'ubicacion_especifica',
                    'responsable_custodio', 'proceso', 'uso_sihos',
                    'uso_office_basico', 'software_especializado', 'descripcion_software',
                    'funcion_principal', 'criticidad', 'confidencialidad',
                    'horario_uso', 'estado_operativo', 'observaciones_tecnicas', 
                    'periodicidad_mtto', 'responsable_mtto'
                ]
                
                col = 4
                for field in naranja_fields:
                    value = ws.cell(row=target_row, column=col).value or ''
                    self.equipment_data[field] = value
                    
                    # Cargar en widgets con verificación
                    if field in self.manual_widgets:
                        try:
                            widget = self.manual_widgets[field]
                            if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                                if isinstance(widget, ctk.CTkEntry):
                                    widget.delete(0, "end")
                                    widget.insert(0, value)
                                elif isinstance(widget, ctk.CTkComboBox):
                                    widget.set(value)
                        except:
                            pass  # Si falla, continuar con el siguiente
                    
                    col += 1
                
                wb.close()
                
                self.equipo_update_code = codigo
                self.equipo_update_row = target_row
                
                # CAMBIAR TÍTULO A MODO ACTUALIZACIÓN (con verificación)
                if hasattr(self, 'equipo_form_frame'):
                    try:
                        if hasattr(self.equipo_form_frame, 'winfo_exists') and self.equipo_form_frame.winfo_exists():
                            self.equipo_form_frame.configure(
                                label_text=f"🔄 ACTUALIZANDO EQUIPO - Código: {codigo}"
                            )
                    except:
                        pass
                
                # CAMBIAR TEXTO DEL BOTÓN (con verificación)
                if hasattr(self, 'btn_save_equipo'):
                    try:
                        if hasattr(self.btn_save_equipo, 'winfo_exists') and self.btn_save_equipo.winfo_exists():
                            self.btn_save_equipo.configure(text="🔄 ACTUALIZAR EQUIPO")
                    except:
                        pass
                
                dialog.destroy()
                
                if messagebox.askyesno(
                    "Confirmar Actualización",
                    f"⚠️ ¿Estás seguro de actualizar {codigo}?\n\n"
                    f"Los datos actuales se han cargado.\n"
                    f"Modifica los campos necesarios y presiona ACTUALIZAR EQUIPO."
                ):
                    messagebox.showinfo("Listo", f"✅ Datos cargados de {codigo}\n\nModifica los campos y presiona ACTUALIZAR EQUIPO.")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al buscar:\n{e}")
        
        btn_buscar = ctk.CTkButton(
            dialog,
            text="🔍 BUSCAR Y CARGAR",
            command=buscar_y_cargar,
            font=("Segoe UI", 13, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            height=40
        )
        btn_buscar.pack(pady=10)
        entry_codigo.bind("<Return>", lambda e: buscar_y_cargar())
    
    def reset_after_update_equipos(self):
        """Reseteo completo después de actualizar un equipo (volver al estado inicial)."""
        # 1. Limpiar variables de modo actualización
        self.equipo_update_row = None
        self.equipo_update_code = None
        
        # 2. Restaurar título al SIGUIENTE equipo nuevo (con verificación)
        next_code = f"EQC-{self.current_row-1:04d}"
        if hasattr(self, 'equipo_form_frame'):
            try:
                if hasattr(self.equipo_form_frame, 'winfo_exists') and self.equipo_form_frame.winfo_exists():
                    self.equipo_form_frame.configure(
                        label_text=f"📝 DATOS MANUALES - Equipo #{self.current_row-1} (Código {next_code})"
                    )
            except:
                pass
        
        # 3. Restaurar BOTÓN a estado normal (con verificación)
        if hasattr(self, 'btn_save_equipo'):
            try:
                if hasattr(self.btn_save_equipo, 'winfo_exists') and self.btn_save_equipo.winfo_exists():
                    self.btn_save_equipo.configure(text="💾 GUARDAR NUEVO (Solo Datos Manuales)")
            except:
                pass
        
        # 4. Limpiar TODOS los datos
        self.equipment_data = {}
        self.verde_data = {}
        self.azul_data = {}
        
        # 5. Limpiar TODOS los widgets (con verificación)
        for key, widget in self.manual_widgets.items():
            try:
                if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                    if isinstance(widget, ctk.CTkEntry):
                        widget.delete(0, "end")
                    elif isinstance(widget, ctk.CTkComboBox):
                        widget.set("")
            except:
                pass
    
    def save_equipo_update(self):
        """Guardar actualización de equipo de cómputo (solo datos manuales)."""
        try:
            wb = load_workbook(self.excel_path)
            ws = wb["Equipos de Cómputo"]
            
            row = self.equipo_update_row
            codigo = self.equipo_update_code
            
            # Actualizar NARANJAS (columnas 4-27)
            col = 4
            naranja_fields = [
                'tipo_equipo', 'area_servicio', 'ubicacion_especifica',
                'responsable_custodio', 'proceso', 'uso_sihos',
                'uso_office_basico', 'software_especializado', 'descripcion_software',
                'funcion_principal','horario_uso', 'estado_operativo', 'observaciones_tecnicas',
                'periodicidad_mtto', 'responsable_mtto'
            ]
            
            # Leer de widgets directamente con verificación (igual que save_equipo_manual_only)
            for field in naranja_fields:
                value = ''
                if field in self.manual_widgets:
                    try:
                        widget = self.manual_widgets[field]
                        if hasattr(widget, 'winfo_exists') and widget.winfo_exists():
                            if isinstance(widget, ctk.CTkEntry):
                                value = widget.get()
                            elif isinstance(widget, ctk.CTkComboBox):
                                value = widget.get()
                    except:
                        pass
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            wb.save(self.excel_path)
            wb.close()
            
            messagebox.showinfo("Éxito", f"✅ Equipo {codigo} actualizado correctamente")
            
            # Reseteo completo usando función unificada
            self.reset_after_update_equipos()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al actualizar:\n{e}")
    
    def show_completion_message(self):
        """Mostrar mensaje de equipo completado."""
        consecutive = self.current_row - 1
        code = f"EQC-{consecutive:04d}"
        nombre = self.verde_data.get('nombre_equipo', 'N/A')
        area = self.equipment_data.get('area_servicio', 'N/A')
        
        # Ventana de completado
        completion_window = ctk.CTkToplevel(self.root)
        completion_window.title("Equipo Completado")
        completion_window.geometry("500x450")
        completion_window.transient(self.root)
        completion_window.grab_set()
        
        # Centrar
        completion_window.update_idletasks()
        x = (completion_window.winfo_screenwidth() // 2) - 250
        y = (completion_window.winfo_screenheight() // 2) - 225
        completion_window.geometry(f"500x450+{x}+{y}")
        
        # Icono de éxito
        success_label = ctk.CTkLabel(
            completion_window,
            text="✅",
            font=("Arial", 60)
        )
        success_label.pack(pady=20)
        
        title = ctk.CTkLabel(
            completion_window,
            text="EQUIPO COMPLETADO Y GUARDADO",
            font=("Arial", 16, "bold"),
            text_color=COLOR_VERDE_HOSPITAL
        )
        title.pack(pady=10)
        
        info_frame = ctk.CTkFrame(completion_window, fg_color="transparent")
        info_frame.pack(pady=20)
        
        info_text = f"""
N° Consecutivo: {consecutive}
Código: {code}
Nombre: {nombre}
Área: {area}

✓ Datos manuales: Guardados
✓ Datos automáticos: Guardados
✓ Datos mixtos: Guardados
✓ Excel actualizado correctamente

Total: 56 columnas completas
        """
        
        info = ctk.CTkLabel(
            info_frame,
            text=info_text,
            font=("Arial", 11),
            justify="left"
        )
        info.pack()
        
        # Instrucciones
        instructions = ctk.CTkLabel(
            completion_window,
            text="Cierra la aplicación para hacer salida segura del USB\ny proceder al siguiente equipo.",
            font=("Arial", 11),
            text_color="gray"
        )
        instructions.pack(pady=10)
        
        # Botones
        btn_frame = ctk.CTkFrame(completion_window, fg_color="transparent")
        btn_frame.pack(pady=20)
        
        btn_next = ctk.CTkButton(
            btn_frame,
            text="➡️ Siguiente Equipo",
            command=lambda: self.next_equipment(completion_window),
            fg_color=COLOR_VERDE_HOSPITAL,
            width=200
        )
        btn_next.pack(side="left", padx=10)
        
        btn_close = ctk.CTkButton(
            btn_frame,
            text="❌ Cerrar",
            command=self.root.quit,
            fg_color=COLOR_ERROR,
            hover_color="#A02828",
            width=200
        )
        btn_close.pack(side="left", padx=10)
    
    def next_equipment(self, completion_window):
        """Ir a siguiente equipo."""
        completion_window.destroy()
        self.current_row += 1
        
        # Limpiar datos
        self.equipment_data = {}
        self.verde_data = {}
        self.azul_data = {}
        
        # Mostrar nuevo formulario
        self.show_manual_form()


    
    # ========================================================================
    # FORMULARIOS DE OTROS TIPOS DE INVENTARIO
    # ========================================================================
    
    def create_impresoras_form(self, parent_tab):
        """Formulario para Impresoras y Escáneres."""
        scroll = ctk.CTkScrollableFrame(
            parent_tab,
            fg_color="#FAFAFA",
            label_fg_color=COLOR_VERDE_HOSPITAL,
            label_text_color="white",
            label_font=("Segoe UI", 15, "bold")
        )
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Detectar siguiente código automáticamente
        next_code = self.detect_next_code("Impresoras y Escáneres", "IMP")
        scroll.configure(label_text=f"🖨️ IMPRESORAS Y ESCÁNERES - Código: {next_code}")
        
        # Widgets para almacenar referencias
        self.imp_widgets = {}
        self.imp_next_code = next_code  # Guardar código para usar al guardar
        self.imp_scroll = scroll  # Guardar referencia al scroll para actualizar título
        
        # Campos
        fields = [
            ("Código Equipo Asignado *", "codigo_asignado", "entry"),
            ("Tipo *", "tipo", "combobox", TIPOS_IMPRESORA),
            ("Marca *", "marca", "combobox", MARCAS_IMPRESORA),
            ("Modelo", "modelo", "entry"),
            ("Serial", "serial", "entry"),
            ("Área / Servicio *", "area", "combobox", AREAS_SERVICIO),
            ("Ubicación Específica", "ubicacion", "entry"),
            ("Función *", "funcion", "combobox", FUNCIONES_IMPRESORA),
            ("Dirección IP", "ip", "entry"),
            ("Estado Operativo *", "estado", "combobox", ESTADOS_IMPRESORA),
            ("Observaciones", "observaciones", "entry"),
        ]
        
        for field_data in fields:
            if len(field_data) == 4:
                label, key, field_type, options = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, options)
            else:
                label, key, field_type = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, None)
            self.imp_widgets[key] = widget
        
        # Frame para botones
        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.pack(pady=30)
        
        # Botón guardar nuevo - Referencia global
        self.btn_save_imp = ctk.CTkButton(
            btn_frame,
            text="💾 GUARDAR NUEVO",
            command=self.save_impresora,
            font=("Segoe UI", 14, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5039",
            height=50,
            width=250
        )
        self.btn_save_imp.pack(side="left", padx=10)
        
        # Botón actualizar existente
        btn_update = ctk.CTkButton(
            btn_frame,
            text="🔄 ACTUALIZAR EXISTENTE",
            command=self.update_impresora,
            font=("Segoe UI", 14, "bold"),
            fg_color="#2196F3",
            hover_color="#1976D2",
            height=50,
            width=250
        )
        btn_update.pack(side="left", padx=10)
    
    def save_impresora(self):
        """Guardar impresora en Excel."""
        if not self.excel_path:
            messagebox.showerror("Error", "No hay Excel cargado")
            return
        
        try:
            wb = load_workbook(self.excel_path)
            
            # Verificar que la hoja existe
            if "Impresoras y Escáneres" not in wb.sheetnames:
                wb.close()
                messagebox.showerror("Error", "La hoja 'Impresoras y Escáneres' no existe en el Excel.\n\nCrea esta hoja primero.")
                return
            
            ws = wb["Impresoras y Escáneres"]
            
            # Verificar si es actualización o nuevo registro
            if hasattr(self, 'imp_update_row') and self.imp_update_row:
                # MODO ACTUALIZACIÓN
                row = self.imp_update_row
                codigo = self.imp_update_code
                
                # Actualizar datos en la fila existente (NO modificar columnas 1 y 2)
                ws.cell(row=row, column=3, value=self.imp_widgets["codigo_asignado"].get())
                ws.cell(row=row, column=4, value=self.imp_widgets["tipo"].get())
                ws.cell(row=row, column=5, value=self.imp_widgets["marca"].get())
                ws.cell(row=row, column=6, value=self.imp_widgets["modelo"].get())
                ws.cell(row=row, column=7, value=self.imp_widgets["serial"].get())
                ws.cell(row=row, column=8, value=self.imp_widgets["area"].get())
                ws.cell(row=row, column=9, value=self.imp_widgets["ubicacion"].get())
                ws.cell(row=row, column=10, value=self.imp_widgets["funcion"].get())
                ws.cell(row=row, column=11, value=self.imp_widgets["ip"].get())
                ws.cell(row=row, column=12, value=self.imp_widgets["estado"].get())
                ws.cell(row=row, column=15, value=self.imp_widgets["observaciones"].get())
                
                wb.save(self.excel_path)
                wb.close()
                
                messagebox.showinfo("Éxito", f"✅ Impresora {codigo} actualizada correctamente")
                
                # Limpiar modo actualización
                self.imp_update_row = None
                self.imp_update_code = None
                
                # Volver a título normal
                next_code = self.detect_next_code("Impresoras y Escáneres", "IMP")
                self.imp_next_code = next_code
                self.imp_scroll.configure(label_text=f"🖨️ IMPRESORAS Y ESCÁNERES - Código: {next_code}")
                
                # Restaurar texto del botón
                if hasattr(self, 'btn_save_imp'):
                    self.btn_save_imp.configure(text="💾 GUARDAR NUEVO")
                
                # Limpiar todos los campos
                for key, widget in self.imp_widgets.items():
                    if isinstance(widget, ctk.CTkEntry):
                        widget.delete(0, "end")
                    elif isinstance(widget, ctk.CTkComboBox):
                        widget.set("")
                
            else:
                # MODO GUARDAR NUEVO
                # Buscar el último consecutivo real
                last_consecutive = 0
                for row in range(2, 200):
                    value = ws.cell(row=row, column=1).value
                    if value is not None:
                        try:
                            consecutivo = int(value)
                            if consecutivo > last_consecutive:
                                last_consecutive = consecutivo
                        except:
                            pass
                
                # Siguiente consecutivo
                next_consecutive = last_consecutive + 1
                
                # Buscar primera fila vacía
                next_row = 2
                for row in range(2, 200):
                    if ws.cell(row=row, column=2).value is None:
                        next_row = row
                        break
                
                # Guardar datos
                ws.cell(row=next_row, column=1, value=next_consecutive)
                ws.cell(row=next_row, column=2, value=f"IMP-{next_consecutive:04d}")
                ws.cell(row=next_row, column=3, value=self.imp_widgets["codigo_asignado"].get())
                ws.cell(row=next_row, column=4, value=self.imp_widgets["tipo"].get())
                ws.cell(row=next_row, column=5, value=self.imp_widgets["marca"].get())
                ws.cell(row=next_row, column=6, value=self.imp_widgets["modelo"].get())
                ws.cell(row=next_row, column=7, value=self.imp_widgets["serial"].get())
                ws.cell(row=next_row, column=8, value=self.imp_widgets["area"].get())
                ws.cell(row=next_row, column=9, value=self.imp_widgets["ubicacion"].get())
                ws.cell(row=next_row, column=10, value=self.imp_widgets["funcion"].get())
                ws.cell(row=next_row, column=11, value=self.imp_widgets["ip"].get())
                ws.cell(row=next_row, column=12, value=self.imp_widgets["estado"].get())
                ws.cell(row=next_row, column=15, value=self.imp_widgets["observaciones"].get())
                
                wb.save(self.excel_path)
                wb.close()
                
                messagebox.showinfo("Éxito", f"✅ Impresora guardada: IMP-{next_consecutive:04d}")
                
                # Detectar siguiente código y actualizar título
                next_code = self.detect_next_code("Impresoras y Escáneres", "IMP")
                self.imp_next_code = next_code
                self.imp_scroll.configure(label_text=f"🖨️ IMPRESORAS Y ESCÁNERES - Código: {next_code}")
                
                # Limpiar campos selectivamente (mantener área)
                campos_a_mantener = ['area']
                
                for key, widget in self.imp_widgets.items():
                    if key not in campos_a_mantener:
                        if isinstance(widget, ctk.CTkEntry):
                            widget.delete(0, "end")
                        elif isinstance(widget, ctk.CTkComboBox):
                            widget.set("")
                    
        except Exception as e:
            messagebox.showerror("Error", f"❌ Error al guardar impresora:\n\n{str(e)}\n\nVerifica que la hoja 'Impresoras y Escáneres' existe.")
            import traceback
            traceback.print_exc()
    
    def update_impresora(self):
        """Actualizar impresora existente en Excel."""
        if not self.excel_path:
            messagebox.showerror("Error", "No hay Excel cargado")
            return
        
        # Ventana para pedir código
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Actualizar Impresora")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Centrar
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 200
        y = (dialog.winfo_screenheight() // 2) - 100
        dialog.geometry(f"400x200+{x}+{y}")
        
        ctk.CTkLabel(
            dialog,
            text="Ingresa el código de la impresora a actualizar:",
            font=("Segoe UI", 13)
        ).pack(pady=20)
        
        entry_codigo = ctk.CTkEntry(
            dialog,
            width=200,
            height=40,
            font=("Segoe UI", 12),
            placeholder_text="Ej: IMP-0026"
        )
        entry_codigo.pack(pady=10)
        entry_codigo.focus()
        
        def buscar_y_cargar():
            codigo = entry_codigo.get().strip().upper()
            if not codigo:
                messagebox.showerror("Error", "Debes ingresar un código")
                return
            
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["Impresoras y Escáneres"]
                
                # Buscar el código en la columna 2
                found = False
                target_row = None
                
                for row in range(2, 200):
                    cell_value = ws.cell(row=row, column=2).value
                    if cell_value and cell_value.upper() == codigo:
                        found = True
                        target_row = row
                        break
                
                if not found:
                    wb.close()
                    messagebox.showerror("Error", f"No se encontró el código {codigo}")
                    return
                
                # Cargar datos en los widgets
                self.imp_widgets["codigo_asignado"].delete(0, "end")
                self.imp_widgets["codigo_asignado"].insert(0, ws.cell(row=target_row, column=3).value or "")
                
                self.imp_widgets["tipo"].set(ws.cell(row=target_row, column=4).value or "")
                self.imp_widgets["marca"].set(ws.cell(row=target_row, column=5).value or "")
                
                self.imp_widgets["modelo"].delete(0, "end")
                self.imp_widgets["modelo"].insert(0, ws.cell(row=target_row, column=6).value or "")
                
                self.imp_widgets["serial"].delete(0, "end")
                self.imp_widgets["serial"].insert(0, ws.cell(row=target_row, column=7).value or "")
                
                self.imp_widgets["area"].set(ws.cell(row=target_row, column=8).value or "")
                
                self.imp_widgets["ubicacion"].delete(0, "end")
                self.imp_widgets["ubicacion"].insert(0, ws.cell(row=target_row, column=9).value or "")
                
                self.imp_widgets["funcion"].set(ws.cell(row=target_row, column=10).value or "")
                
                self.imp_widgets["ip"].delete(0, "end")
                self.imp_widgets["ip"].insert(0, ws.cell(row=target_row, column=11).value or "")
                
                self.imp_widgets["estado"].set(ws.cell(row=target_row, column=12).value or "")
                
                self.imp_widgets["observaciones"].delete(0, "end")
                self.imp_widgets["observaciones"].insert(0, ws.cell(row=target_row, column=15).value or "")
                
                wb.close()
                
                # Guardar código y fila para actualizar
                self.imp_update_code = codigo
                self.imp_update_row = target_row
                
                # CAMBIAR TÍTULO A MODO ACTUALIZACIÓN
                self.imp_scroll.configure(label_text=f"🔄 ACTUALIZANDO IMPRESORA - Código: {codigo}")
                
                # CAMBIAR TEXTO DEL BOTÓN
                if hasattr(self, 'btn_save_imp'):
                    self.btn_save_imp.configure(text="🔄 ACTUALIZAR IMPRESORA")
                
                dialog.destroy()
                
                # Confirmar
                if messagebox.askyesno(
                    "Confirmar Actualización",
                    f"⚠️ ¿Estás seguro de actualizar {codigo}?\n\n"
                    f"Los datos actuales se han cargado.\n"
                    f"Modifica los campos necesarios y presiona GUARDAR NUEVO."
                ):
                    messagebox.showinfo("Listo", f"✅ Datos cargados de {codigo}\n\nModifica los campos y presiona GUARDAR NUEVO.")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al buscar:\n{e}")
        
        btn_buscar = ctk.CTkButton(
            dialog,
            text="🔍 BUSCAR Y CARGAR",
            command=buscar_y_cargar,
            font=("Segoe UI", 13, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            height=40
        )
        btn_buscar.pack(pady=10)
        
        # Enter para buscar
        entry_codigo.bind("<Return>", lambda e: buscar_y_cargar())
    
    def create_perifericos_form(self, parent_tab):
        """Formulario para Periféricos."""
        scroll = ctk.CTkScrollableFrame(
            parent_tab,
            fg_color="#FAFAFA",
            label_fg_color=COLOR_VERDE_HOSPITAL,
            label_text_color="white",
            label_font=("Segoe UI", 15, "bold")
        )
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Detectar siguiente código automáticamente
        next_code = self.detect_next_code("Periféricos", "PER")
        scroll.configure(label_text=f"🖱️ PERIFÉRICOS - Código: {next_code}")
        
        self.per_widgets = {}
        self.per_next_code = next_code
        self.per_scroll = scroll  # Guardar referencia al scroll
        
        fields = [
            ("Código Equipo Asignado *", "codigo_asignado", "entry"),
            ("Tipo *", "tipo", "combobox", TIPOS_PERIFERICO),
            ("Marca *", "marca", "combobox", MARCAS_PERIFERICO),
            ("Modelo", "modelo", "entry"),
            ("Serial", "serial", "entry"),
            ("Área / Servicio *", "area", "combobox", AREAS_SERVICIO),
            ("Estado Operativo *", "estado", "combobox", ESTADOS_PERIFERICO),
            ("Observaciones", "observaciones", "entry"),
        ]
        
        for field_data in fields:
            if len(field_data) == 4:
                label, key, field_type, options = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, options)
            else:
                label, key, field_type = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, None)
            self.per_widgets[key] = widget
        
        # Frame para botones
        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.pack(pady=30)
        
        # Botón guardar nuevo - Referencia global
        self.btn_save_per = ctk.CTkButton(
            btn_frame,
            text="💾 GUARDAR NUEVO",
            command=self.save_periferico,
            font=("Segoe UI", 14, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5039",
            height=50,
            width=250
        )
        self.btn_save_per.pack(side="left", padx=10)
        
        # Botón actualizar existente
        btn_update = ctk.CTkButton(
            btn_frame,
            text="🔄 ACTUALIZAR EXISTENTE",
            command=self.update_periferico,
            font=("Segoe UI", 14, "bold"),
            fg_color="#2196F3",
            hover_color="#1976D2",
            height=50,
            width=250
        )
        btn_update.pack(side="left", padx=10)
    
    def save_periferico(self):
        """Guardar periférico en Excel."""
        if not self.excel_path:
            messagebox.showerror("Error", "No hay Excel cargado")
            return
        
        try:
            wb = load_workbook(self.excel_path)
            
            # Verificar que la hoja existe
            if "Periféricos" not in wb.sheetnames:
                wb.close()
                messagebox.showerror("Error", "La hoja 'Periféricos' no existe en el Excel. Crea esta hoja primero.")
                return
            
            ws = wb["Periféricos"]
            
            # Verificar si es actualización o nuevo registro
            if hasattr(self, 'per_update_row') and self.per_update_row:
                # MODO ACTUALIZACIÓN
                row = self.per_update_row
                codigo = self.per_update_code
                
                # Actualizar datos en la fila existente
                ws.cell(row=row, column=3, value=self.per_widgets["codigo_asignado"].get())
                ws.cell(row=row, column=4, value=self.per_widgets["tipo"].get())
                ws.cell(row=row, column=5, value=self.per_widgets["marca"].get())
                ws.cell(row=row, column=6, value=self.per_widgets["modelo"].get())
                ws.cell(row=row, column=7, value=self.per_widgets["serial"].get())
                ws.cell(row=row, column=8, value=self.per_widgets["area"].get())
                ws.cell(row=row, column=9, value=self.per_widgets["estado"].get())
                ws.cell(row=row, column=11, value=self.per_widgets["observaciones"].get())
                
                wb.save(self.excel_path)
                wb.close()
                
                messagebox.showinfo("Éxito", f"✅ Periférico {codigo} actualizado correctamente")
                
                # Limpiar modo actualización
                self.per_update_row = None
                self.per_update_code = None
                
                # Volver a título normal
                next_code = self.detect_next_code("Periféricos", "PER")
                self.per_next_code = next_code
                self.per_scroll.configure(label_text=f"🖱️ PERIFÉRICOS - Código: {next_code}")
                
                # Restaurar texto del botón
                if hasattr(self, 'btn_save_per'):
                    self.btn_save_per.configure(text="💾 GUARDAR NUEVO")
                
                # Limpiar todos los campos
                for key, widget in self.per_widgets.items():
                    if isinstance(widget, ctk.CTkEntry):
                        widget.delete(0, "end")
                    elif isinstance(widget, ctk.CTkComboBox):
                        widget.set("")
                
            else:
                # MODO GUARDAR NUEVO
                # Buscar el último consecutivo real
                last_consecutive = 0
                for row in range(2, 200):
                    value = ws.cell(row=row, column=1).value
                    if value is not None:
                        try:
                            consecutivo = int(value)
                            if consecutivo > last_consecutive:
                                last_consecutive = consecutivo
                        except:
                            pass
                
                # Siguiente consecutivo
                next_consecutive = last_consecutive + 1
                
                # Buscar primera fila vacía
                next_row = 2
                for row in range(2, 200):
                    if ws.cell(row=row, column=2).value is None:
                        next_row = row
                        break
                
                ws.cell(row=next_row, column=1, value=next_consecutive)
                ws.cell(row=next_row, column=2, value=f"PER-{next_consecutive:04d}")
                ws.cell(row=next_row, column=3, value=self.per_widgets["codigo_asignado"].get())
                ws.cell(row=next_row, column=4, value=self.per_widgets["tipo"].get())
                ws.cell(row=next_row, column=5, value=self.per_widgets["marca"].get())
                ws.cell(row=next_row, column=6, value=self.per_widgets["modelo"].get())
                ws.cell(row=next_row, column=7, value=self.per_widgets["serial"].get())
                ws.cell(row=next_row, column=8, value=self.per_widgets["area"].get())
                ws.cell(row=next_row, column=9, value=self.per_widgets["estado"].get())
                ws.cell(row=next_row, column=11, value=self.per_widgets["observaciones"].get())
                
                wb.save(self.excel_path)
                wb.close()
                
                messagebox.showinfo("Éxito", f"✅ Periférico guardado: PER-{next_consecutive:04d}")
                
                # Detectar siguiente código y actualizar título
                next_code = self.detect_next_code("Periféricos", "PER")
                self.per_next_code = next_code
                self.per_scroll.configure(label_text=f"🖱️ PERIFÉRICOS - Código: {next_code}")
                
                # Limpiar campos selectivamente (mantener área)
                campos_a_mantener = ['area']
                
                for key, widget in self.per_widgets.items():
                    if key not in campos_a_mantener:
                        if isinstance(widget, ctk.CTkEntry):
                            widget.delete(0, "end")
                        elif isinstance(widget, ctk.CTkComboBox):
                            widget.set("")
                    
        except Exception as e:
            messagebox.showerror("Error", f"❌ Error al guardar periférico:\n\n{str(e)}\n\nVerifica que la hoja 'Periféricos' existe.")
            import traceback
            traceback.print_exc()
    
    def update_periferico(self):
        """Actualizar periférico existente en Excel."""
        if not self.excel_path:
            messagebox.showerror("Error", "No hay Excel cargado")
            return
        
        # Ventana para pedir código
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Actualizar Periférico")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 200
        y = (dialog.winfo_screenheight() // 2) - 100
        dialog.geometry(f"400x200+{x}+{y}")
        
        ctk.CTkLabel(
            dialog,
            text="Ingresa el código del periférico a actualizar:",
            font=("Segoe UI", 13)
        ).pack(pady=20)
        
        entry_codigo = ctk.CTkEntry(
            dialog,
            width=200,
            height=40,
            font=("Segoe UI", 12),
            placeholder_text="Ej: PER-0026"
        )
        entry_codigo.pack(pady=10)
        entry_codigo.focus()
        
        def buscar_y_cargar():
            codigo = entry_codigo.get().strip().upper()
            if not codigo:
                messagebox.showerror("Error", "Debes ingresar un código")
                return
            
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["Periféricos"]
                
                found = False
                target_row = None
                
                for row in range(2, 200):
                    cell_value = ws.cell(row=row, column=2).value
                    if cell_value and cell_value.upper() == codigo:
                        found = True
                        target_row = row
                        break
                
                if not found:
                    wb.close()
                    messagebox.showerror("Error", f"No se encontró el código {codigo}")
                    return
                
                # Cargar datos
                self.per_widgets["codigo_asignado"].delete(0, "end")
                self.per_widgets["codigo_asignado"].insert(0, ws.cell(row=target_row, column=3).value or "")
                
                self.per_widgets["tipo"].set(ws.cell(row=target_row, column=4).value or "")
                self.per_widgets["marca"].set(ws.cell(row=target_row, column=5).value or "")
                
                self.per_widgets["modelo"].delete(0, "end")
                self.per_widgets["modelo"].insert(0, ws.cell(row=target_row, column=6).value or "")
                
                self.per_widgets["serial"].delete(0, "end")
                self.per_widgets["serial"].insert(0, ws.cell(row=target_row, column=7).value or "")
                
                self.per_widgets["area"].set(ws.cell(row=target_row, column=8).value or "")
                self.per_widgets["estado"].set(ws.cell(row=target_row, column=9).value or "")
                
                self.per_widgets["observaciones"].delete(0, "end")
                self.per_widgets["observaciones"].insert(0, ws.cell(row=target_row, column=11).value or "")
                
                wb.close()
                
                self.per_update_code = codigo
                self.per_update_row = target_row
                
                # CAMBIAR TÍTULO A MODO ACTUALIZACIÓN
                self.per_scroll.configure(label_text=f"🔄 ACTUALIZANDO PERIFÉRICO - Código: {codigo}")
                
                # CAMBIAR TEXTO DEL BOTÓN
                if hasattr(self, 'btn_save_per'):
                    self.btn_save_per.configure(text="🔄 ACTUALIZAR PERIFÉRICO")
                
                dialog.destroy()
                
                if messagebox.askyesno(
                    "Confirmar Actualización",
                    f"⚠️ ¿Estás seguro de actualizar {codigo}?\n\n"
                    f"Los datos actuales se han cargado.\n"
                    f"Modifica los campos necesarios y presiona GUARDAR NUEVO."
                ):
                    messagebox.showinfo("Listo", f"✅ Datos cargados de {codigo}\n\nModifica los campos y presiona GUARDAR NUEVO.")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al buscar:\n{e}")
        
        btn_buscar = ctk.CTkButton(
            dialog,
            text="🔍 BUSCAR Y CARGAR",
            command=buscar_y_cargar,
            font=("Segoe UI", 13, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            height=40
        )
        btn_buscar.pack(pady=10)
        entry_codigo.bind("<Return>", lambda e: buscar_y_cargar())
    
    def create_red_form(self, parent_tab):
        """Formulario para Equipos de Red."""
        scroll = ctk.CTkScrollableFrame(
            parent_tab,
            fg_color="#FAFAFA",
            label_fg_color=COLOR_VERDE_HOSPITAL,
            label_text_color="white",
            label_font=("Segoe UI", 15, "bold")
        )
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Detectar siguiente código automáticamente
        next_code = self.detect_next_code("Equipos de Red", "RED")
        scroll.configure(label_text=f"🌐 EQUIPOS DE RED - Código: {next_code}")
        
        self.red_widgets = {}
        self.red_next_code = next_code
        self.red_scroll = scroll  # Guardar referencia al scroll
        
        fields = [
            ("Tipo *", "tipo", "combobox", TIPOS_EQUIPO_RED),
            ("Marca *", "marca", "combobox", MARCAS_RED),
            ("Modelo", "modelo", "entry"),
            ("Serial", "serial", "entry"),
            ("Dirección IP *", "ip", "entry"),
            ("Puertos Totales", "puertos", "entry"),
            ("Ubicación *", "ubicacion", "combobox", UBICACIONES_RED),
            ("Área / Servicio", "area", "combobox", AREAS_SERVICIO),
            ("Estado Operativo *", "estado", "combobox", ESTADOS_RED),
            ("Observaciones", "observaciones", "entry"),
        ]
        
        for field_data in fields:
            if len(field_data) == 4:
                label, key, field_type, options = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, options)
            else:
                label, key, field_type = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, None)
            self.red_widgets[key] = widget
        
        # Frame para botones
        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.pack(pady=30)
        
        # Botón guardar nuevo - Referencia global
        self.btn_save_red = ctk.CTkButton(
            btn_frame,
            text="💾 GUARDAR NUEVO",
            command=self.save_red,
            font=("Segoe UI", 14, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5039",
            height=50,
            width=250
        )
        self.btn_save_red.pack(side="left", padx=10)
        
        # Botón actualizar existente
        btn_update = ctk.CTkButton(
            btn_frame,
            text="🔄 ACTUALIZAR EXISTENTE",
            command=self.update_red,
            font=("Segoe UI", 14, "bold"),
            fg_color="#2196F3",
            hover_color="#1976D2",
            height=50,
            width=250
        )
        btn_update.pack(side="left", padx=10)
    
    def save_red(self):
        """Guardar equipo de red en Excel."""
        if not self.excel_path:
            messagebox.showerror("Error", "No hay Excel cargado")
            return
        
        try:
            wb = load_workbook(self.excel_path)
            
            # Verificar que la hoja existe
            if "Equipos de Red" not in wb.sheetnames:
                wb.close()
                messagebox.showerror("Error", "La hoja 'Equipos de Red' no existe en el Excel.\n\nCrea esta hoja primero.")
                return
            
            ws = wb["Equipos de Red"]
            
            # Verificar si es actualización o nuevo registro
            if hasattr(self, 'red_update_row') and self.red_update_row:
                # MODO ACTUALIZACIÓN
                row = self.red_update_row
                codigo = self.red_update_code
                
                # Actualizar datos en la fila existente (NO modificar columnas 1 y 2)
                ws.cell(row=row, column=3, value=self.red_widgets["tipo"].get())
                ws.cell(row=row, column=4, value=self.red_widgets["marca"].get())
                ws.cell(row=row, column=5, value=self.red_widgets["modelo"].get())
                ws.cell(row=row, column=6, value=self.red_widgets["serial"].get())
                ws.cell(row=row, column=7, value=self.red_widgets["ip"].get())
                ws.cell(row=row, column=8, value=self.red_widgets["puertos"].get())
                ws.cell(row=row, column=9, value=self.red_widgets["ubicacion"].get())
                ws.cell(row=row, column=10, value=self.red_widgets["area"].get())
                ws.cell(row=row, column=11, value=self.red_widgets["estado"].get())
                ws.cell(row=row, column=14, value=self.red_widgets["observaciones"].get())
                
                wb.save(self.excel_path)
                wb.close()
                
                messagebox.showinfo("Éxito", f"✅ Equipo de red {codigo} actualizado correctamente")
                
                # Limpiar modo actualización
                self.red_update_row = None
                self.red_update_code = None
                
                # Volver a título normal
                next_code = self.detect_next_code("Equipos de Red", "RED")
                self.red_next_code = next_code
                self.red_scroll.configure(label_text=f"🌐 EQUIPOS DE RED - Código: {next_code}")
                
                # Restaurar texto del botón
                if hasattr(self, 'btn_save_red'):
                    self.btn_save_red.configure(text="💾 GUARDAR NUEVO")
                
                # Limpiar todos los campos
                for key, widget in self.red_widgets.items():
                    if isinstance(widget, ctk.CTkEntry):
                        widget.delete(0, "end")
                    elif isinstance(widget, ctk.CTkComboBox):
                        widget.set("")
                
            else:
                # MODO GUARDAR NUEVO
                # Buscar el último consecutivo real
                last_consecutive = 0
                for row in range(2, 100):
                    value = ws.cell(row=row, column=1).value
                    if value is not None:
                        try:
                            consecutivo = int(value)
                            if consecutivo > last_consecutive:
                                last_consecutive = consecutivo
                        except:
                            pass
                
                # Siguiente consecutivo
                next_consecutive = last_consecutive + 1
                
                # Buscar primera fila vacía
                next_row = 2
                for row in range(2, 100):
                    if ws.cell(row=row, column=2).value is None:
                        next_row = row
                        break
                
                ws.cell(row=next_row, column=1, value=next_consecutive)
                ws.cell(row=next_row, column=2, value=f"RED-{next_consecutive:04d}")
                ws.cell(row=next_row, column=3, value=self.red_widgets["tipo"].get())
                ws.cell(row=next_row, column=4, value=self.red_widgets["marca"].get())
                ws.cell(row=next_row, column=5, value=self.red_widgets["modelo"].get())
                ws.cell(row=next_row, column=6, value=self.red_widgets["serial"].get())
                ws.cell(row=next_row, column=7, value=self.red_widgets["ip"].get())
                ws.cell(row=next_row, column=8, value=self.red_widgets["puertos"].get())
                ws.cell(row=next_row, column=9, value=self.red_widgets["ubicacion"].get())
                ws.cell(row=next_row, column=10, value=self.red_widgets["area"].get())
                ws.cell(row=next_row, column=11, value=self.red_widgets["estado"].get())
                ws.cell(row=next_row, column=14, value=self.red_widgets["observaciones"].get())
                
                wb.save(self.excel_path)
                wb.close()
                
                messagebox.showinfo("Éxito", f"✅ Equipo de red guardado: RED-{next_consecutive:04d}")
                
                # Detectar siguiente código y actualizar título
                next_code = self.detect_next_code("Equipos de Red", "RED")
                self.red_next_code = next_code
                self.red_scroll.configure(label_text=f"🌐 EQUIPOS DE RED - Código: {next_code}")
                
                # Limpiar campos selectivamente (mantener área y ubicación)
                campos_a_mantener = ['area', 'ubicacion']
                
                for key, widget in self.red_widgets.items():
                    if key not in campos_a_mantener:
                        if isinstance(widget, ctk.CTkEntry):
                            widget.delete(0, "end")
                        elif isinstance(widget, ctk.CTkComboBox):
                            widget.set("")
                    
        except Exception as e:
            messagebox.showerror("Error", f"❌ Error al guardar equipo de red:\n\n{str(e)}\n\nVerifica que la hoja 'Equipos de Red' existe.")
            import traceback
            traceback.print_exc()
    
    def update_red(self):
        """Actualizar equipo de red existente en Excel."""
        if not self.excel_path:
            messagebox.showerror("Error", "No hay Excel cargado")
            return
        
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("Actualizar Equipo de Red")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - 200
        y = (dialog.winfo_screenheight() // 2) - 100
        dialog.geometry(f"400x200+{x}+{y}")
        
        ctk.CTkLabel(
            dialog,
            text="Ingresa el código del equipo de red a actualizar:",
            font=("Segoe UI", 13)
        ).pack(pady=20)
        
        entry_codigo = ctk.CTkEntry(
            dialog,
            width=200,
            height=40,
            font=("Segoe UI", 12),
            placeholder_text="Ej: RED-0026"
        )
        entry_codigo.pack(pady=10)
        entry_codigo.focus()
        
        def buscar_y_cargar():
            codigo = entry_codigo.get().strip().upper()
            if not codigo:
                messagebox.showerror("Error", "Debes ingresar un código")
                return
            
            try:
                wb = load_workbook(self.excel_path)
                ws = wb["Equipos de Red"]
                
                found = False
                target_row = None
                
                for row in range(2, 100):
                    cell_value = ws.cell(row=row, column=2).value
                    if cell_value and cell_value.upper() == codigo:
                        found = True
                        target_row = row
                        break
                
                if not found:
                    wb.close()
                    messagebox.showerror("Error", f"No se encontró el código {codigo}")
                    return
                
                # Cargar datos
                self.red_widgets["tipo"].set(ws.cell(row=target_row, column=3).value or "")
                self.red_widgets["marca"].set(ws.cell(row=target_row, column=4).value or "")
                
                self.red_widgets["modelo"].delete(0, "end")
                self.red_widgets["modelo"].insert(0, ws.cell(row=target_row, column=5).value or "")
                
                self.red_widgets["serial"].delete(0, "end")
                self.red_widgets["serial"].insert(0, ws.cell(row=target_row, column=6).value or "")
                
                self.red_widgets["ip"].delete(0, "end")
                self.red_widgets["ip"].insert(0, ws.cell(row=target_row, column=7).value or "")
                
                self.red_widgets["puertos"].delete(0, "end")
                self.red_widgets["puertos"].insert(0, ws.cell(row=target_row, column=8).value or "")
                
                self.red_widgets["ubicacion"].set(ws.cell(row=target_row, column=9).value or "")
                self.red_widgets["area"].set(ws.cell(row=target_row, column=10).value or "")
                self.red_widgets["estado"].set(ws.cell(row=target_row, column=11).value or "")
                
                self.red_widgets["observaciones"].delete(0, "end")
                self.red_widgets["observaciones"].insert(0, ws.cell(row=target_row, column=14).value or "")
                
                wb.close()
                
                self.red_update_code = codigo
                self.red_update_row = target_row
                
                # CAMBIAR TÍTULO A MODO ACTUALIZACIÓN
                self.red_scroll.configure(label_text=f"🔄 ACTUALIZANDO EQUIPO DE RED - Código: {codigo}")
                
                # CAMBIAR TEXTO DEL BOTÓN
                if hasattr(self, 'btn_save_red'):
                    self.btn_save_red.configure(text="🔄 ACTUALIZAR EQUIPO DE RED")
                
                dialog.destroy()
                
                if messagebox.askyesno(
                    "Confirmar Actualización",
                    f"⚠️ ¿Estás seguro de actualizar {codigo}?\n\n"
                    f"Los datos actuales se han cargado.\n"
                    f"Modifica los campos necesarios y presiona GUARDAR NUEVO."
                ):
                    messagebox.showinfo("Listo", f"✅ Datos cargados de {codigo}\n\nModifica los campos y presiona GUARDAR NUEVO.")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al buscar:\n{e}")
        
        btn_buscar = ctk.CTkButton(
            dialog,
            text="🔍 BUSCAR Y CARGAR",
            command=buscar_y_cargar,
            font=("Segoe UI", 13, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            height=40
        )
        btn_buscar.pack(pady=10)
        entry_codigo.bind("<Return>", lambda e: buscar_y_cargar())
    
    def create_mantenimientos_form(self, parent_tab):
        """Formulario para Mantenimientos."""
        scroll = ctk.CTkScrollableFrame(
            parent_tab,
            fg_color="#FAFAFA",
            label_fg_color=COLOR_VERDE_HOSPITAL,
            label_text_color="white",
            label_font=("Segoe UI", 15, "bold")
        )
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Detectar siguiente consecutivo
        next_consecutive = self.detect_next_consecutive_mantenimiento()
        scroll.configure(label_text=f"🔧 MANTENIMIENTOS - Registro #{next_consecutive}")
        
        self.mtt_widgets = {}
        self.mtt_scroll = scroll  # Guardar referencia
        self.mtt_next_consecutive = next_consecutive
        
        fields = [
            ("Código Equipo *", "codigo_equipo", "entry"),
            ("Tipo Mantenimiento *", "tipo", "combobox", TIPOS_MANTENIMIENTO_MTTO),
            ("Técnico Responsable *", "tecnico", "combobox", TECNICOS_RESPONSABLES),
            ("Descripción Actividades *", "descripcion", "combobox", ACTIVIDADES_MANTENIMIENTO),
            ("Repuestos/Insumos", "repuestos", "entry"),
            ("Estado Post-Mtto *", "estado_post", "combobox", ESTADO_POST_MTTO),
            ("Observaciones", "observaciones", "entry"),
        ]

        self.create_date_field_centered(scroll, "Fecha Mantenimiento *", "fecha_mtto")
        self.create_date_field_centered(scroll, "Próximo Mantenimiento", "proximo")

        for field_data in fields:
            if len(field_data) == 4:
                label, key, field_type, options = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, options)
            else:
                label, key, field_type = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, None)
            self.mtt_widgets[key] = widget
        
        btn_save = ctk.CTkButton(
            scroll,
            text="💾 GUARDAR MANTENIMIENTO",
            command=self.save_mantenimiento,
            font=("Segoe UI", 14, "bold"),
            fg_color=COLOR_VERDE_HOSPITAL,
            hover_color="#1F5039",
            height=50
        )
        btn_save.pack(pady=30)
    
    def save_mantenimiento(self):
        """Guardar mantenimiento en Excel."""
        if not self.excel_path:
            messagebox.showerror("Error", "No hay Excel cargado")
            return
        
        try:
            wb = load_workbook(self.excel_path)
            ws = wb["Mantenimientos"]
            
            next_row = 2
            for row in range(2, 500):
                if ws.cell(row=row, column=1).value is None:
                    next_row = row
                    break
            
            consecutive = next_row - 1
            
            ws.cell(row=next_row, column=1, value=consecutive)
            ws.cell(row=next_row, column=2, value=self.mtt_widgets["codigo_equipo"].get())
            ws.cell(row=next_row, column=3, value=self.get_date_value(self.mtt_widgets["fecha_mtto"]))
            ws.cell(row=next_row, column=4, value=self.mtt_widgets["tipo"].get())
            ws.cell(row=next_row, column=5, value=self.mtt_widgets["tecnico"].get())
            ws.cell(row=next_row, column=6, value=self.mtt_widgets["descripcion"].get())
            ws.cell(row=next_row, column=7, value=self.mtt_widgets["repuestos"].get())
            ws.cell(row=next_row, column=8, value=self.mtt_widgets["estado_post"].get())
            ws.cell(row=next_row, column=9, value=self.mtt_widgets["proximo"].get())
            ws.cell(row=next_row, column=10, value=self.mtt_widgets["observaciones"].get())
            
            wb.save(self.excel_path)
            wb.close()
            
            messagebox.showinfo("Éxito", f"✅ Mantenimiento registrado #{consecutive}")
            
            # Actualizar título para siguiente registro
            next_consecutive = self.detect_next_consecutive_mantenimiento()
            self.mtt_next_consecutive = next_consecutive
            self.mtt_scroll.configure(label_text=f"🔧 MANTENIMIENTOS - Registro #{next_consecutive}")
            
            # Limpiar campos selectivamente (mantener técnico)
            campos_a_mantener = ['tecnico']
            
            for key, widget in self.mtt_widgets.items():
                if key not in campos_a_mantener:
                    if isinstance(widget, ctk.CTkEntry):
                        widget.delete(0, "end")
                    elif isinstance(widget, ctk.CTkComboBox):
                        widget.set("")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar:\n{e}")
    
    def create_baja_form(self, parent_tab):
        """Formulario para Equipos Dados de Baja."""
        scroll = ctk.CTkScrollableFrame(
            parent_tab,
            fg_color="#FAFAFA",
            label_fg_color=COLOR_VERDE_HOSPITAL,
            label_text_color="white",
            label_font=("Segoe UI", 15, "bold")
        )
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Detectar siguiente número de baja
        next_baja = self.detect_next_baja()
        scroll.configure(label_text=f"📦 EQUIPOS DADOS DE BAJA - Baja #{next_baja}")
        
        self.baja_widgets = {}
        self.baja_scroll = scroll  # Guardar referencia
        self.baja_next = next_baja
        
        fields = [
            ("Código Original *", "codigo_original", "entry"),
            ("Tipo *", "tipo", "entry"),
            ("Marca", "marca", "entry"),
            ("Modelo", "modelo", "entry"),
            ("Serial", "serial", "entry"),
            ("Motivo Baja *", "motivo", "combobox", MOTIVOS_BAJA),
            ("Destino *", "destino", "combobox", DESTINOS_BAJA),
            ("Responsable Baja *", "responsable", "combobox", RESPONSABLES_BAJA),
            ("Observaciones", "observaciones", "entry"),
        ]

        self.create_date_field_centered(scroll, "Fecha de Baja *", "fecha_baja")
        
        for field_data in fields:
            if len(field_data) == 4:
                label, key, field_type, options = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, options)
            else:
                label, key, field_type = field_data
                widget = self.create_form_field_centered(scroll, label, key, field_type, None)
            self.baja_widgets[key] = widget
        
        # Botón para buscar y autocompletar
        btn_search = ctk.CTkButton(
            scroll,
            text="🔍 BUSCAR Y AUTOCOMPLETAR DESDE INVENTARIO",
            command=self.buscar_equipo_baja,
            font=("Segoe UI", 13, "bold"),
            fg_color="#2196F3",
            hover_color="#1976D2",
            height=45
        )
        btn_search.pack(pady=15, padx=20, fill="x")
        
        # Separador
        separator = ctk.CTkFrame(scroll, height=2, fg_color="#E0E0E0")
        separator.pack(fill="x", padx=20, pady=10)
        
        btn_save = ctk.CTkButton(
            scroll,
            text="💾 REGISTRAR BAJA",
            command=self.save_baja,
            font=("Segoe UI", 14, "bold"),
            fg_color="#DC3545",
            hover_color="#A02828",
            height=50
        )
        btn_save.pack(pady=30)
    
    def buscar_equipo_baja(self):
        """Buscar equipo en inventarios y autocompletar datos."""
        codigo = self.baja_widgets["codigo_original"].get().strip().upper()
        
        if not codigo:
            messagebox.showerror("Error", "Primero ingresa el código del equipo en el campo 'Código Original'")
            return
        
        try:
            wb = load_workbook(self.excel_path)
            
            # Determinar en qué hoja buscar según el prefijo
            if codigo.startswith("EQC-"):
                ws_name = "Equipos de Cómputo"
                col_codigo = 2
            elif codigo.startswith("IMP-"):
                ws_name = "Impresoras y Escáneres"
                col_codigo = 2
            elif codigo.startswith("PER-"):
                ws_name = "Periféricos"
                col_codigo = 2
            elif codigo.startswith("RED-"):
                ws_name = "Equipos de Red"
                col_codigo = 2
            else:
                wb.close()
                messagebox.showerror("Error", "Código no válido. Usa: EQC-XXXX, IMP-XXX, PER-XXX, RED-XXX")
                return
            
            ws = wb[ws_name]
            found = False
            target_row = None
            
            # Buscar código
            for row in range(2, 500):
                cell_value = ws.cell(row=row, column=col_codigo).value
                if cell_value and cell_value.upper() == codigo:
                    found = True
                    target_row = row
                    break
            
            if not found:
                wb.close()
                messagebox.showerror("Error", f"No se encontró el código {codigo} en {ws_name}")
                return
            
            # Autocompletar según el tipo
            if codigo.startswith("EQC-"):
                # Equipos de Cómputo
                tipo = ws.cell(row=target_row, column=4).value or "Computador"  # Tipo equipo
                marca = ws.cell(row=target_row, column=28).value or ""  # Marca (verde)
                modelo = ws.cell(row=target_row, column=29).value or ""  # Modelo (verde)
                serial = ws.cell(row=target_row, column=30).value or ""  # Serial (verde)
                
            elif codigo.startswith("IMP-"):
                # Impresoras
                tipo = ws.cell(row=target_row, column=4).value or "Impresora"
                marca = ws.cell(row=target_row, column=5).value or ""
                modelo = ws.cell(row=target_row, column=6).value or ""
                serial = ws.cell(row=target_row, column=7).value or ""
                
            elif codigo.startswith("PER-"):
                # Periféricos
                tipo = ws.cell(row=target_row, column=4).value or "Periférico"
                marca = ws.cell(row=target_row, column=5).value or ""
                modelo = ws.cell(row=target_row, column=6).value or ""
                serial = ws.cell(row=target_row, column=7).value or ""
                
            elif codigo.startswith("RED-"):
                # Equipos de Red
                tipo = ws.cell(row=target_row, column=3).value or "Equipo de Red"
                marca = ws.cell(row=target_row, column=4).value or ""
                modelo = ws.cell(row=target_row, column=5).value or ""
                serial = ws.cell(row=target_row, column=6).value or ""
            
            wb.close()
            
            # Cargar datos en los widgets
            self.baja_widgets["tipo"].delete(0, "end")
            self.baja_widgets["tipo"].insert(0, tipo)
            
            self.baja_widgets["marca"].delete(0, "end")
            self.baja_widgets["marca"].insert(0, marca)
            
            self.baja_widgets["modelo"].delete(0, "end")
            self.baja_widgets["modelo"].insert(0, modelo)
            
            self.baja_widgets["serial"].delete(0, "end")
            self.baja_widgets["serial"].insert(0, serial)
            
            # Guardar información para actualizar después
            self.baja_origen_sheet = ws_name
            self.baja_origen_row = target_row
            
            messagebox.showinfo("Éxito", f"✅ Datos cargados de {codigo}\n\nCompleta los campos de baja y guarda.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al buscar equipo:\n{e}")
    
    def save_baja(self):
        """Guardar equipo dado de baja en Excel y actualizar estado en inventario original."""
        if not self.excel_path:
            messagebox.showerror("Error", "No hay Excel cargado")
            return
        
        try:
            wb = load_workbook(self.excel_path)
            ws_baja = wb["Equipos Dados de Baja"]
            
            next_row = 2
            for row in range(2, 200):
                if ws_baja.cell(row=row, column=1).value is None:
                    next_row = row
                    break
            
            codigo = self.baja_widgets["codigo_original"].get()
            
            # Guardar en hoja de Dados de Baja
            ws_baja.cell(row=next_row, column=1, value=codigo)
            ws_baja.cell(row=next_row, column=2, value=self.baja_widgets["tipo"].get())
            ws_baja.cell(row=next_row, column=3, value=self.baja_widgets["marca"].get())
            ws_baja.cell(row=next_row, column=4, value=self.baja_widgets["modelo"].get())
            ws_baja.cell(row=next_row, column=5, value=self.baja_widgets["serial"].get())
            ws_baja.cell(row=next_row, column=6, value=self.get_date_value(self.baja_widgets["fecha_baja"]))
            ws_baja.cell(row=next_row, column=7, value=self.baja_widgets["motivo"].get())
            ws_baja.cell(row=next_row, column=8, value=self.baja_widgets["destino"].get())
            ws_baja.cell(row=next_row, column=9, value=self.baja_widgets["responsable"].get())
            ws_baja.cell(row=next_row, column=10, value=self.baja_widgets["observaciones"].get())
            
            # Actualizar estado en inventario original (si fue autocompletado)
            if hasattr(self, 'baja_origen_sheet') and hasattr(self, 'baja_origen_row'):
                ws_origen = wb[self.baja_origen_sheet]
                
                # Actualizar estado operativo según el tipo
                if self.baja_origen_sheet == "Equipos de Cómputo":
                    # Columna 15 = Estado Operativo
                    ws_origen.cell(row=self.baja_origen_row, column=15, value="DADO DE BAJA")
                elif self.baja_origen_sheet == "Impresoras y Escáneres":
                    # Columna 12 = Estado
                    ws_origen.cell(row=self.baja_origen_row, column=12, value="DADO DE BAJA")
                elif self.baja_origen_sheet == "Periféricos":
                    # Columna 9 = Estado
                    ws_origen.cell(row=self.baja_origen_row, column=9, value="DADO DE BAJA")
                elif self.baja_origen_sheet == "Equipos de Red":
                    # Columna 11 = Estado
                    ws_origen.cell(row=self.baja_origen_row, column=11, value="DADO DE BAJA")
                
                # Limpiar referencias
                delattr(self, 'baja_origen_sheet')
                delattr(self, 'baja_origen_row')
            
            wb.save(self.excel_path)
            wb.close()
            
            messagebox.showinfo("Éxito", 
                f"✅ Baja registrada: {codigo}\n\n"
                f"• Agregado a 'Equipos Dados de Baja'\n"
                f"• Estado actualizado a 'DADO DE BAJA' en inventario original")
            
            # Actualizar título para siguiente registro
            next_baja = self.detect_next_baja()
            self.baja_next = next_baja
            self.baja_scroll.configure(label_text=f"📦 EQUIPOS DADOS DE BAJA - Baja #{next_baja}")
            
            # Limpiar campos selectivamente
            campos_a_mantener = ['responsable', 'motivo', 'destino']
            
            for key, widget in self.baja_widgets.items():
                if key not in campos_a_mantener:
                    if isinstance(widget, ctk.CTkEntry):
                        widget.delete(0, "end")
                    elif isinstance(widget, ctk.CTkComboBox):
                        widget.set("")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar:\n{e}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    """Función principal."""
    root = ctk.CTk()
    app = InventoryManagerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
